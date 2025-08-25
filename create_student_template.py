import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def create_student_template():
    """创建学生信息导入模板Excel文件"""
    # 创建工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "学生信息"
    
    # 设置标题行
    headers = ["学生姓名", "学号", "班级", "手机号", "邮箱"]
    
    # 写入标题行
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        # 设置标题样式
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 添加示例数据
    sample_data = [
        ["张三", "2024001", "计算机科学与技术1班", "13800138001", "zhangsan@example.com"],
        ["李四", "2024002", "计算机科学与技术1班", "13800138002", "lisi@example.com"],
        ["王五", "2024003", "计算机科学与技术2班", "13800138003", "wangwu@example.com"],
        ["赵六", "2024004", "计算机科学与技术2班", "13800138004", "zhaoliu@example.com"],
        ["钱七", "2024005", "软件工程1班", "13800138005", "qianqi@example.com"]
    ]
    
    # 写入示例数据
    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # 设置列宽
    column_widths = [15, 12, 25, 15, 25]
    for col_idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    
    # 保存文件
    filename = "templates/学生信息导入模板.xlsx"
    wb.save(filename)
    print(f"学生信息导入模板已创建：{filename}")
    
    return filename

if __name__ == "__main__":
    create_student_template()