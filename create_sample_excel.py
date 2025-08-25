from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

def create_sample_excel():
    """创建学生信息导入示例Excel文件"""
    
    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "学生信息导入模板"
    
    # 设置标题行
    headers = ['学生姓名', '学号', '班级', '手机号', '邮箱']
    ws.append(headers)
    
    # 设置标题行样式
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # 添加示例数据
    sample_data = [
        ['张三', '2021001', '计算机科学与技术1班', '13800138001', 'zhangsan@example.com'],
        ['李四', '2021002', '计算机科学与技术1班', '13800138002', 'lisi@example.com'],
        ['王五', '2021003', '计算机科学与技术2班', '13800138003', 'wangwu@example.com'],
        ['赵六', '2021004', '计算机科学与技术2班', '13800138004', 'zhaoliu@example.com'],
        ['钱七', '2021005', '软件工程1班', '13800138005', 'qianqi@example.com'],
        ['孙八', '2021006', '软件工程1班', '13800138006', 'sunba@example.com'],
        ['周九', '2021007', '软件工程2班', '13800138007', 'zhoujiu@example.com'],
        ['吴十', '2021008', '软件工程2班', '13800138008', 'wushi@example.com'],
        ['郑十一', '2021009', '数据科学与大数据技术1班', '13800138009', 'zhengshiyi@example.com'],
        ['王十二', '2021010', '数据科学与大数据技术1班', '13800138010', 'wangshier@example.com']
    ]
    
    for row_data in sample_data:
        ws.append(row_data)
    
    # 调整列宽
    column_widths = [15, 15, 25, 15, 25]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = width
    
    # 保存文件
    filename = '学生信息导入模板.xlsx'
    wb.save(filename)
    print(f"示例Excel文件已创建: {filename}")
    
    return filename

if __name__ == '__main__':
    create_sample_excel()