import os
import sqlite3
import json
from flask import Flask, request, jsonify, send_file, render_template
from flask_cors import CORS
from werkzeug.security import generate_password_hash
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import io

# 数据库配置
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
PROJ_ROOT = os.path.dirname(BASE_DIR)
DB_DIR = os.path.join(PROJ_ROOT, 'DB')
DB_PATH = os.path.join(DB_DIR, 'app.db')

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

# 学生信息管理类
class StudentManager:
    def __init__(self):
        self.ensure_student_table()
    
    def ensure_student_table(self):
        """确保学生表存在"""
        with get_db() as conn:
            # 检查users表是否已有学生相关字段
            cursor = conn.cursor()
            cursor.execute("PRAGMA table_info(users)")
            columns = [row[1] for row in cursor.fetchall()]
            
            # 如果缺少字段，添加它们
            if 'class_name' not in columns:
                cursor.execute('ALTER TABLE users ADD COLUMN class_name TEXT')
            if 'phone' not in columns:
                cursor.execute('ALTER TABLE users ADD COLUMN phone TEXT')
            if 'email' not in columns:
                cursor.execute('ALTER TABLE users ADD COLUMN email TEXT')
            if 'status' not in columns:
                cursor.execute('ALTER TABLE users ADD COLUMN status TEXT DEFAULT "active"')
            if 'created_at' not in columns:
                cursor.execute('ALTER TABLE users ADD COLUMN created_at DATETIME')
    
    def import_from_excel(self, file_path_or_stream):
        """从Excel文件导入学生信息"""
        try:
            # 加载Excel文件
            wb = load_workbook(filename=file_path_or_stream, data_only=True)
            ws = wb.active
            
            imported_count = 0
            updated_count = 0
            error_rows = []
            
            with get_db() as conn:
                cursor = conn.cursor()
                
                # 遍历Excel行（跳过标题行）
                for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
                    if row_idx == 1:  # 跳过标题行
                        continue
                    
                    try:
                        # 解析Excel列：学生姓名、学号、班级、手机、邮箱
                        name = str(row[0]).strip() if row and row[0] is not None else None
                        student_id = str(row[1]).strip() if row and len(row) > 1 and row[1] is not None else None
                        class_name = str(row[2]).strip() if row and len(row) > 2 and row[2] is not None else None
                        phone = str(row[3]).strip() if row and len(row) > 3 and row[3] is not None else None
                        email = str(row[4]).strip() if row and len(row) > 4 and row[4] is not None else None
                        
                        # 验证必填字段
                        if not name or not student_id:
                            error_rows.append(f"第{row_idx}行：姓名和学号为必填项")
                            continue
                        
                        # 检查学号是否已存在
                        existing = cursor.execute(
                            'SELECT id, name FROM users WHERE student_id = ?', 
                            (student_id,)
                        ).fetchone()
                        
                        if existing:
                            # 更新现有记录
                            cursor.execute('''
                                UPDATE users SET 
                                    name = ?, 
                                    class_name = ?, 
                                    phone = ?, 
                                    email = ?
                                WHERE student_id = ?
                            ''', (name, class_name, phone, email, student_id))
                            updated_count += 1
                        else:
                            # 插入新记录
                            cursor.execute('''
                                INSERT INTO users (
                                    student_id, name, class_name, phone, email, 
                                    role, password_hash, status, created_at
                                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                            ''', (
                                student_id, name, class_name, phone, email,
                                'student', generate_password_hash('123456'), 'active',
                                datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                            ))
                            imported_count += 1
                    
                    except Exception as e:
                        error_rows.append(f"第{row_idx}行：{str(e)}")
                        continue
            
            return {
                'success': True,
                'imported': imported_count,
                'updated': updated_count,
                'errors': error_rows,
                'total_processed': imported_count + updated_count
            }
        
        except Exception as e:
            return {
                'success': False,
                'error': f"文件处理失败：{str(e)}",
                'imported': 0,
                'updated': 0,
                'errors': []
            }
    
    def export_to_excel(self):
        """导出学生信息到Excel"""
        try:
            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "学生信息"
            
            # 设置标题行
            headers = ['学生姓名', '学号', '班级', '手机号', '邮箱', '状态', '创建时间']
            ws.append(headers)
            
            # 设置标题行样式
            header_font = Font(bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center')
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # 获取学生数据
            with get_db() as conn:
                cursor = conn.cursor()
                students = cursor.execute('''
                    SELECT name, student_id, class_name, phone, email, status, created_at
                    FROM users 
                    WHERE role = 'student'
                    ORDER BY student_id
                ''').fetchall()
                
                # 添加数据行
                for student in students:
                    ws.append([
                        student['name'],
                        student['student_id'],
                        student['class_name'] or '',
                        student['phone'] or '',
                        student['email'] or '',
                        student['status'] or 'active',
                        student['created_at'] or ''
                    ])
            
            # 调整列宽
            column_widths = [15, 15, 15, 15, 25, 10, 20]
            for col_num, width in enumerate(column_widths, 1):
                ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = width
            
            # 保存到内存
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            return output
        
        except Exception as e:
            raise Exception(f"导出失败：{str(e)}")
    
    def get_students_list(self, page=1, per_page=20, search=None):
        """获取学生列表（支持分页和搜索）"""
        try:
            offset = (page - 1) * per_page
            
            with get_db() as conn:
                cursor = conn.cursor()
                
                # 构建查询条件
                where_clause = "WHERE role = 'student'"
                params = []
                
                if search:
                    where_clause += " AND (name LIKE ? OR student_id LIKE ? OR class_name LIKE ?)"
                    search_param = f"%{search}%"
                    params.extend([search_param, search_param, search_param])
                
                # 获取总数
                total_query = f"SELECT COUNT(*) FROM users {where_clause}"
                total = cursor.execute(total_query, params).fetchone()[0]
                
                # 获取分页数据
                data_query = f'''
                    SELECT id, name, student_id, class_name, phone, email, status, created_at
                    FROM users {where_clause}
                    ORDER BY student_id
                    LIMIT ? OFFSET ?
                '''
                params.extend([per_page, offset])
                
                students = cursor.execute(data_query, params).fetchall()
                
                return {
                    'students': [dict(student) for student in students],
                    'total': total,
                    'page': page,
                    'per_page': per_page,
                    'pages': (total + per_page - 1) // per_page
                }
        
        except Exception as e:
            raise Exception(f"获取学生列表失败：{str(e)}")
    
    def delete_student(self, student_id):
        """删除学生"""
        try:
            with get_db() as conn:
                cursor = conn.cursor()
                
                # 检查学生是否存在
                student = cursor.execute(
                    'SELECT id FROM users WHERE student_id = ? AND role = "student"',
                    (student_id,)
                ).fetchone()
                
                if not student:
                    return {'success': False, 'error': '学生不存在'}
                
                # 删除学生（软删除，设置状态为deleted）
                cursor.execute(
                    'UPDATE users SET status = "deleted" WHERE student_id = ?',
                    (student_id,)
                )
                
                return {'success': True, 'message': '学生删除成功'}
        
        except Exception as e:
            return {'success': False, 'error': f'删除失败：{str(e)}'}
    
    def get_student_stats(self):
        """获取学生统计信息"""
        try:
            with get_db() as conn:
                cursor = conn.cursor()
                
                # 总学生数
                total_students = cursor.execute(
                    'SELECT COUNT(*) FROM users WHERE role = "student" AND status != "deleted"'
                ).fetchone()[0]
                
                # 按班级统计
                class_stats = cursor.execute('''
                    SELECT class_name, COUNT(*) as count
                    FROM users 
                    WHERE role = 'student' AND status != 'deleted' AND class_name IS NOT NULL
                    GROUP BY class_name
                    ORDER BY count DESC
                ''').fetchall()
                
                # 最近注册的学生
                recent_students = cursor.execute('''
                    SELECT name, student_id, created_at
                    FROM users 
                    WHERE role = 'student' AND status != 'deleted'
                    ORDER BY created_at DESC
                    LIMIT 5
                ''').fetchall()
                
                return {
                    'total_students': total_students,
                    'class_stats': [dict(row) for row in class_stats],
                    'recent_students': [dict(row) for row in recent_students]
                }
        
        except Exception as e:
            raise Exception(f"获取统计信息失败：{str(e)}")

# 创建Flask应用
def create_student_app():
    app = Flask(__name__, template_folder='../templates')
    CORS(app)
    
    student_manager = StudentManager()
    
    @app.route('/api/students/import', methods=['POST'])
    def import_students():
        """导入学生信息"""
        if 'file' not in request.files:
            return jsonify({'error': '缺少文件'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': '未选择文件'}), 400
        
        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({'error': '文件格式不正确，请上传Excel文件'}), 400
        
        try:
            result = student_manager.import_from_excel(file)
            return jsonify(result)
        except Exception as e:
            return jsonify({'error': f'导入失败：{str(e)}'}), 500
    
    @app.route('/api/students/export', methods=['GET'])
    def export_students():
        """导出学生信息"""
        try:
            output = student_manager.export_to_excel()
            filename = f"学生信息_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            return send_file(
                output,
                as_attachment=True,
                download_name=filename,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        except Exception as e:
            return jsonify({'error': f'导出失败：{str(e)}'}), 500
    
    @app.route('/api/students', methods=['GET'])
    def get_students():
        """获取学生列表"""
        try:
            page = int(request.args.get('page', 1))
            per_page = int(request.args.get('per_page', 20))
            search = request.args.get('search', '').strip()
            
            result = student_manager.get_students_list(
                page=page, 
                per_page=per_page, 
                search=search if search else None
            )
            return jsonify(result)
        except Exception as e:
            return jsonify({'error': f'获取学生列表失败：{str(e)}'}), 500
    
    @app.route('/api/students/<student_id>', methods=['DELETE'])
    def delete_student(student_id):
        """删除学生"""
        try:
            result = student_manager.delete_student(student_id)
            return jsonify(result)
        except Exception as e:
            return jsonify({'error': f'删除失败：{str(e)}'}), 500
    
    @app.route('/api/students/stats', methods=['GET'])
    def get_student_stats():
        """获取学生统计信息"""
        try:
            result = student_manager.get_student_stats()
            return jsonify(result)
        except Exception as e:
            return jsonify({'error': f'获取统计信息失败：{str(e)}'}), 500
    
    @app.route('/')
    def index():
        return render_template('student_manager.html')
    
    return app

if __name__ == '__main__':
    app = create_student_app()
    app.run(host='0.0.0.0', port=92, debug=True)