import os, json, sqlite3, random, re, time, secrets
from flask import Flask, request, jsonify
from flask_cors import CORS
from werkzeug.security import generate_password_hash, check_password_hash

try:
    from openai import OpenAI
except Exception:
    OpenAI = None

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
PROJ_ROOT = os.path.dirname(BASE_DIR)
DB_DIR = os.path.join(PROJ_ROOT, 'DB')
os.makedirs(DB_DIR, exist_ok=True)
DB_PATH = os.path.join(DB_DIR, 'app.db')

app = Flask(__name__)
CORS(app)

# ---------- DB Utils ----------

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

SCHEMA_SQL = '''
CREATE TABLE IF NOT EXISTS modules (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  title TEXT NOT NULL,
  description TEXT,
  ord INTEGER
);
CREATE TABLE IF NOT EXISTS topics (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  module_id INTEGER NOT NULL,
  title TEXT NOT NULL,
  ord INTEGER,
  FOREIGN KEY(module_id) REFERENCES modules(id)
);
CREATE TABLE IF NOT EXISTS contents (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  topic_id INTEGER NOT NULL,
  data TEXT NOT NULL,
  FOREIGN KEY(topic_id) REFERENCES topics(id)
);
CREATE TABLE IF NOT EXISTS exam_sets (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  name TEXT NOT NULL
);
CREATE TABLE IF NOT EXISTS questions (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  exam_id INTEGER NOT NULL,
  qtype TEXT NOT NULL,
  prompt TEXT NOT NULL,
  options TEXT,
  answer TEXT NOT NULL,
  score INTEGER NOT NULL,
  ord INTEGER,
  knowledge_ref TEXT,
  FOREIGN KEY(exam_id) REFERENCES exam_sets(id)
);
CREATE TABLE IF NOT EXISTS users (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  username TEXT UNIQUE,
  student_id TEXT UNIQUE,
  name TEXT,
  role TEXT NOT NULL DEFAULT 'student',
  password_hash TEXT NOT NULL
);
CREATE TABLE IF NOT EXISTS tokens (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  user_id INTEGER NOT NULL,
  token TEXT NOT NULL,
  created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
  FOREIGN KEY(user_id) REFERENCES users(id)
);
CREATE TABLE IF NOT EXISTS submissions (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  user_id INTEGER NOT NULL,
  exam_id INTEGER NOT NULL,
  score INTEGER NOT NULL,
  total INTEGER NOT NULL,
  rate REAL NOT NULL,
  detail TEXT NOT NULL,
  wrong_qids TEXT,
  suggestions TEXT,
  created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
  FOREIGN KEY(user_id) REFERENCES users(id),
  FOREIGN KEY(exam_id) REFERENCES exam_sets(id)
);
'''

# Ensure column exists for existing DBs

def ensure_column(cursor, table, column):
    info = cursor.execute(f"PRAGMA table_info({table})").fetchall()
    cols = [r[1] for r in info]
    if column not in cols:
        cursor.execute(f"ALTER TABLE {table} ADD COLUMN {column} TEXT")


def init_db():
    with get_db() as c:
        c.executescript(SCHEMA_SQL)
        # migrate knowledge_ref if missing
        try:
            ensure_column(c, 'questions', 'knowledge_ref')
        except Exception:
            pass
        # ensure admin
        row = c.execute("SELECT id FROM users WHERE role='admin' LIMIT 1").fetchone()
        if not row:
            c.execute(
                'INSERT INTO users(username,name,role,password_hash) VALUES(?,?,?,?)',
                ('admin', '管理员', 'admin', generate_password_hash('xv2010wr'))
            )

# ---------- Generators ----------
MODULES_META = [
    (1, 'Linux基础操作与环境熟悉', '掌握Linux常用命令、文件系统、权限管理与环境配置'),
    (2, 'Shell脚本基础语法与简单应用', '变量、条件判断、循环控制与函数定义'),
    (3, 'Shell脚本进阶（文件操作与进程监控）', '文件处理、进程管理、任务调度与系统监控'),
    (4, '基于Docker知识点', '容器技术、镜像管理、网络配置与Docker Compose'),
    (5, '基于Docker的Hadoop伪分布式集群搭建', 'Docker环境下Hadoop集群部署与配置实战'),
    (6, 'Hadoop核心组件', 'HDFS分布式文件系统、YARN资源管理器、MapReduce计算框架详解与实战')
]

IMAGES = [
  'https://images.unsplash.com/photo-1518779578993-ec3579fee39f',
  'https://images.unsplash.com/photo-1519389950473-47ba0277781c',
  'https://images.unsplash.com/photo-1508780709619-79562169bc64'
]

def gen_content(module_name: str, topic_idx: int):
    """生成详细的学习内容，包括理论、代码、案例和练习"""
    
    # 根据模块和知识点生成具体内容
    content_map = {
        'Linux基础操作与环境熟悉': {
            1: {
                'title': 'Linux系统概述与发行版选择',
                'theory': '''Linux是一个开源的类Unix操作系统内核，由Linus Torvalds于1991年创建。\n\n**核心概念：**\n• 内核(Kernel)：系统的核心，管理硬件资源\n• 发行版(Distribution)：基于Linux内核的完整操作系统\n• 开源特性：源代码公开，可自由修改和分发\n\n**主流发行版对比：**\n• Ubuntu：用户友好，适合初学者\n• CentOS/RHEL：企业级稳定性，广泛用于服务器\n• Debian：稳定可靠，包管理优秀\n• Arch Linux：滚动更新，高度可定制\n\n**应用场景：**\n• 服务器运维：Web服务器、数据库服务器\n• 开发环境：编程、测试、部署\n• 大数据平台：Hadoop、Spark集群部署''',
                'code': '''# 查看系统信息\nuname -a                    # 系统内核信息\ncat /etc/os-release        # 发行版信息\nlsb_release -a            # 详细版本信息\n\n# 系统资源查看\nfree -h                   # 内存使用情况\ndf -h                     # 磁盘空间\nlscpu                     # CPU信息\n\n# 进程和服务\nps aux                    # 查看所有进程\ntop                       # 实时进程监控\nsystemctl status sshd     # 查看服务状态''',
                'case': '''**案例：搭建开发环境**\n\n场景：为大数据学习搭建Ubuntu虚拟机环境\n\n步骤：\n1. 下载Ubuntu 20.04 LTS镜像\n2. 使用VirtualBox创建虚拟机（4GB内存，50GB硬盘）\n3. 安装系统，配置用户账户\n4. 更新软件源：sudo apt update && sudo apt upgrade\n5. 安装开发工具：sudo apt install vim git curl wget\n\n**实际应用：**\n• 企业通常选择CentOS作为生产环境\n• 开发团队使用Ubuntu进行日常开发\n• 容器化部署多使用Alpine Linux（轻量级）'''
            },
            2: {
                'title': '文件系统结构与目录导航',
                'theory': '''Linux采用树形目录结构，所有文件和目录都从根目录(/)开始。\n\n**重要目录说明：**\n• /：根目录，所有文件的起点\n• /home：用户主目录\n• /etc：系统配置文件\n• /var：可变数据文件（日志、缓存）\n• /usr：用户程序和数据\n• /bin, /sbin：系统命令\n• /tmp：临时文件\n• /opt：第三方软件安装目录\n\n**路径概念：**\n• 绝对路径：从根目录开始，如/home/user/file.txt\n• 相对路径：从当前目录开始，如./file.txt或../parent/\n• 特殊符号：~（用户主目录）、.（当前目录）、..（上级目录）''',
                'code': '''# 目录导航命令\npwd                       # 显示当前目录\nls                        # 列出文件\nls -la                    # 详细列表（包含隐藏文件）\nls -lh                    # 人性化显示文件大小\n\n# 目录切换\ncd /home                  # 切换到指定目录\ncd ~                      # 切换到用户主目录\ncd -                      # 切换到上次访问的目录\ncd ..                     # 切换到上级目录\n\n# 目录操作\nmkdir mydir               # 创建目录\nmkdir -p path/to/dir      # 递归创建目录\nrmdir emptydir            # 删除空目录\nrm -rf dirname            # 强制删除目录及内容''',
                'case': '''**案例：组织项目目录结构**\n\n为大数据项目创建标准目录结构：\n\n```bash\n# 创建项目根目录\nmkdir -p ~/bigdata-project/{data,scripts,config,logs,output}\n\n# 进入项目目录\ncd ~/bigdata-project\n\n# 查看创建的结构\ntree .\n# 或使用 ls -la\n\n# 创建子目录\nmkdir -p data/{raw,processed,external}\nmkdir -p scripts/{etl,analysis,utils}\n```\n\n**目录规划原则：**\n• data/：存放数据文件\n• scripts/：存放脚本代码\n• config/：配置文件\n• logs/：日志文件\n• output/：输出结果'''
            },
            3: {
                'title': '文件操作与权限管理',
                'theory': '''Linux文件权限系统是系统安全的基础，采用用户-组-其他的三级权限模型。\n\n**权限类型：**\n• r (read)：读权限，数值为4\n• w (write)：写权限，数值为2\n• x (execute)：执行权限，数值为1\n\n**权限表示：**\n• 符号表示：rwxrwxrwx（用户-组-其他）\n• 数字表示：755（7=4+2+1, 5=4+1, 5=4+1）\n\n**特殊权限：**\n• SUID：以文件所有者身份执行\n• SGID：以文件所属组身份执行\n• Sticky Bit：只有所有者可删除文件''',
                'code': '''# 文件基本操作\ntouch newfile.txt         # 创建空文件\ncp source.txt dest.txt    # 复制文件\nmv oldname.txt newname.txt # 移动/重命名\nrm filename.txt           # 删除文件\n\n# 权限查看和修改\nls -l filename            # 查看文件权限\nchmod 755 script.sh       # 数字方式设置权限\nchmod u+x script.sh       # 符号方式添加执行权限\nchmod g-w file.txt        # 移除组写权限\n\n# 所有者和组\nchown user:group file.txt # 改变所有者和组\nchown user file.txt       # 只改变所有者\nchgrp group file.txt      # 只改变组''',
                'case': '''**案例：配置Web服务器文件权限**\n\n场景：为Apache Web服务器配置正确的文件权限\n\n```bash\n# 创建网站目录结构\nsudo mkdir -p /var/www/mysite/{public,logs,config}\n\n# 设置目录权限\nsudo chmod 755 /var/www/mysite\nsudo chmod 755 /var/www/mysite/public\nsudo chmod 750 /var/www/mysite/logs\nsudo chmod 750 /var/www/mysite/config\n\n# 设置所有者\nsudo chown -R www-data:www-data /var/www/mysite\n\n# 设置文件权限\nfind /var/www/mysite -type f -exec chmod 644 {} \\;\nfind /var/www/mysite -type d -exec chmod 755 {} \\;\n```\n\n**安全原则：**\n• 最小权限原则：只给必要的权限\n• 定期审查权限设置\n• 敏感文件设置严格权限'''
            },
            4: {
                'title': '进程管理与系统监控',
                'theory': '''Linux进程管理是系统运维的核心技能，涉及进程查看、控制和监控。\n\n**进程概念：**\n• 进程：正在运行的程序实例\n• PID：进程标识符，系统唯一\n• PPID：父进程ID\n• 进程状态：运行(R)、睡眠(S)、僵尸(Z)、停止(T)\n\n**进程优先级：**\n• Nice值：-20到19，值越小优先级越高\n• 实时优先级：0-99，用于实时进程\n\n**信号机制：**\n• SIGTERM(15)：正常终止信号\n• SIGKILL(9)：强制终止信号\n• SIGHUP(1)：挂起信号，常用于重载配置''',
                'code': '''# 进程查看\nps aux                    # 查看所有进程\nps -ef                    # 另一种格式\ntop                       # 实时进程监控\nhtop                      # 增强版top\n\n# 进程控制\nkill 1234                 # 发送TERM信号\nkill -9 1234              # 强制终止\nkillall nginx             # 按名称终止\npkill -f "python app.py"  # 按命令行匹配\n\n# 后台进程\nnohup python app.py &     # 后台运行\njobs                      # 查看作业\nfg %1                     # 前台运行作业1\nbg %1                     # 后台运行作业1''',
                'case': '''**案例：Web服务器进程监控脚本**\n\n创建一个监控Nginx进程的脚本：\n\n```bash\n#!/bin/bash\n# 文件名：monitor_nginx.sh\n\ncheck_nginx() {\n    if pgrep nginx > /dev/null; then\n        echo "$(date): Nginx正在运行"\n        # 检查进程数量\n        process_count=$(pgrep nginx | wc -l)\n        echo "Nginx进程数: $process_count"\n    else\n        echo "$(date): Nginx未运行，尝试启动..."\n        sudo systemctl start nginx\n        if [ $? -eq 0 ]; then\n            echo "Nginx启动成功"\n        else\n            echo "Nginx启动失败"\n        fi\n    fi\n}\n\n# 每30秒检查一次\nwhile true; do\n    check_nginx\n    sleep 30\ndone\n```\n\n**监控要点：**\n• 进程存活检查\n• 资源使用监控\n• 自动重启机制\n• 日志记录'''
            },
            5: {
                'title': '网络配置与服务管理',
                'theory': '''Linux网络配置和服务管理是系统管理员必备技能，涉及网络接口、防火墙、服务管理等。\n\n**网络接口管理：**\n• 网络接口：eth0、wlan0、lo等\n• IP地址配置：静态IP和DHCP\n• 网络工具：ifconfig、ip、netstat、ss\n\n**服务管理（systemd）：**\n• 服务单元：.service文件\n• 服务状态：active、inactive、failed\n• 服务控制：start、stop、restart、reload\n\n**防火墙管理：**\n• iptables：传统防火墙工具\n• ufw：Ubuntu简化防火墙\n• firewalld：CentOS/RHEL防火墙''',
                'code': '''# 网络配置\nip addr show              # 查看网络接口\nip route show             # 查看路由表\nping -c 4 google.com      # 网络连通性测试\nnslookup google.com       # DNS查询\n\n# 网络连接\nnetstat -tulpn            # 查看监听端口\nss -tulpn                 # 现代版netstat\nlsof -i :80               # 查看端口占用\n\n# 服务管理\nsystemctl status nginx    # 查看服务状态\nsystemctl start nginx     # 启动服务\nsystemctl stop nginx      # 停止服务\nsystemctl restart nginx   # 重启服务\nsystemctl enable nginx    # 开机自启\n\n# 防火墙管理\nsudo ufw status           # 查看防火墙状态\nsudo ufw allow 80         # 允许80端口\nsudo ufw deny 22          # 拒绝22端口''',
                'case': '''**案例：配置Web服务器网络环境**\n\n场景：为生产环境配置Nginx服务器\n\n```bash\n#!/bin/bash\n# 配置脚本：setup_webserver.sh\n\n# 1. 配置静态IP（Ubuntu）\nsudo tee /etc/netplan/01-netcfg.yaml > /dev/null <<EOF\nnetwork:\n  version: 2\n  ethernets:\n    eth0:\n      dhcp4: no\n      addresses: [192.168.1.100/24]\n      gateway4: 192.168.1.1\n      nameservers:\n        addresses: [8.8.8.8, 8.8.4.4]\nEOF\n\n# 应用网络配置\nsudo netplan apply\n\n# 2. 安装和配置Nginx\nsudo apt update\nsudo apt install -y nginx\n\n# 3. 配置防火墙\nsudo ufw allow 22          # SSH\nsudo ufw allow 80          # HTTP\nsudo ufw allow 443         # HTTPS\nsudo ufw --force enable\n\n# 4. 启动服务\nsudo systemctl start nginx\nsudo systemctl enable nginx\n\necho "Web服务器配置完成"\n```\n\n**配置要点：**\n• 网络接口配置\n• 防火墙规则设置\n• 服务自动启动\n• 安全加固措施'''
            },
            6: {
                'title': 'VIM编辑器详解与实战应用',
                'theory': '''VIM（Vi IMproved）是Linux系统中最强大的文本编辑器之一，是系统管理员和开发者必备工具。VIM具有高效的编辑能力和丰富的功能。\n\n**VIM的三种模式：**\n• 命令模式（Normal Mode）：默认模式，用于导航和执行命令\n• 插入模式（Insert Mode）：用于输入和编辑文本\n• 可视模式（Visual Mode）：用于选择文本块\n\n**VIM的优势：**\n• 高效的键盘操作，无需鼠标\n• 强大的搜索和替换功能\n• 支持正则表达式\n• 可扩展性强，支持插件\n• 在所有Unix/Linux系统中都可用\n\n**基本概念：**\n• 光标移动：h(左) j(下) k(上) l(右)\n• 文本对象：单词(w)、行(l)、段落(p)\n• 操作符：删除(d)、复制(y)、粘贴(p)\n• 组合操作：操作符+文本对象，如dw(删除单词)''',
                'code': '''# VIM基础操作\n\n# 启动VIM\nvim filename.txt          # 打开文件\nvim +10 file.txt         # 打开文件并跳到第10行\nvim +/pattern file.txt   # 打开文件并搜索pattern\n\n# 模式切换\ni                        # 进入插入模式（光标前）\na                        # 进入插入模式（光标后）\nA                        # 进入插入模式（行尾）\no                        # 新建行并进入插入模式\nO                        # 在上方新建行并进入插入模式\nEsc                      # 返回命令模式\n\n# 光标移动\nh j k l                  # 左下上右\nw                        # 下一个单词开头\nb                        # 上一个单词开头\ne                        # 单词结尾\n0                        # 行首\n$                        # 行尾\ngg                       # 文件开头\nG                        # 文件结尾\n:10                      # 跳到第10行\n\n# 文本编辑\nx                        # 删除当前字符\ndw                       # 删除单词\ndd                       # 删除整行\n3dd                      # 删除3行\nyy                       # 复制当前行\n3yy                      # 复制3行\np                        # 粘贴\nu                        # 撤销\nCtrl+r                   # 重做\n\n# 搜索和替换\n/pattern                 # 向下搜索\n?pattern                 # 向上搜索\nn                        # 下一个匹配\nN                        # 上一个匹配\n:s/old/new/              # 替换当前行第一个\n:s/old/new/g             # 替换当前行所有\n:%s/old/new/g            # 替换全文所有\n:%s/old/new/gc           # 替换全文所有（确认）\n\n# 文件操作\n:w                       # 保存\n:w filename              # 另存为\n:q                       # 退出\n:q!                      # 强制退出不保存\n:wq                      # 保存并退出\n:x                       # 保存并退出（同:wq）''',
                'case': '''**案例：使用VIM进行系统配置文件编辑**\n\n场景：编辑Nginx配置文件并进行批量修改\n\n```bash\n# 1. 备份原配置文件\nsudo cp /etc/nginx/nginx.conf /etc/nginx/nginx.conf.bak\n\n# 2. 使用VIM编辑配置文件\nsudo vim /etc/nginx/nginx.conf\n\n# 在VIM中的操作步骤：\n# - 按 / 搜索 "worker_processes"\n# - 按 n 找到下一个匹配\n# - 按 cw 删除单词并进入插入模式\n# - 输入新值 "auto"\n# - 按 Esc 返回命令模式\n# - 输入 :%s/server_name localhost/server_name example.com/g\n# - 按回车执行全局替换\n# - 输入 :wq 保存并退出\n```\n\n**VIM实用技巧：**\n\n1. **快速编辑多个文件：**\n```bash\nvim file1.txt file2.txt file3.txt\n:next                    # 下一个文件\n:prev                    # 上一个文件\n:buffers                 # 查看缓冲区\n:b2                      # 切换到缓冲区2\n```\n\n2. **使用标记进行快速跳转：**\n```bash\nma                       # 在当前位置设置标记a\n`a                       # 跳转到标记a\n```\n\n3. **分屏操作：**\n```bash\n:split                   # 水平分屏\n:vsplit                  # 垂直分屏\nCtrl+w h/j/k/l          # 在分屏间移动\n```\n\n4. **批量操作：**\n```bash\n# 在多行前添加注释\n1. Ctrl+v 进入可视块模式\n2. j 选择多行\n3. I 进入插入模式\n4. 输入 "# "\n5. Esc 应用到所有选中行\n```\n\n**VIM配置文件示例（~/.vimrc）：**\n```bash\n# 基础配置\nset number               # 显示行号\nset tabstop=4           # Tab宽度\nset shiftwidth=4        # 缩进宽度\nset expandtab           # Tab转空格\nset hlsearch            # 高亮搜索\nset incsearch           # 增量搜索\nset ignorecase          # 忽略大小写\nset smartcase           # 智能大小写\nset autoindent          # 自动缩进\nset syntax=on           # 语法高亮\n\n# 快捷键映射\nmap <F2> :w<CR>         # F2保存\nmap <F3> :q<CR>         # F3退出\n```\n\n**生产环境应用：**\n• 服务器配置文件编辑\n• 日志文件分析和处理\n• 脚本开发和调试\n• 代码审查和修改\n• 批量文本处理任务'''
            },
            7: {
                'title': 'vi编辑器基础操作与实用技巧',
                'theory': '''vi是Unix/Linux系统中最原始和基础的文本编辑器，是VIM的前身。vi编辑器在所有Unix/Linux系统中都默认安装，是系统管理员必须掌握的基本工具。\n\n**vi的特点：**\n• 轻量级，占用资源少\n• 在所有Unix/Linux系统中都可用\n• 纯键盘操作，适合远程管理\n• 启动速度快\n• 功能相对简单但足够实用\n\n**vi的两种模式：**\n• 命令模式（Command Mode）：默认模式，用于导航和执行命令\n• 插入模式（Insert Mode）：用于输入和编辑文本\n\n**与VIM的区别：**\n• vi功能较少，VIM功能更丰富\n• vi语法高亮支持有限\n• vi插件系统不如VIM完善\n• vi在某些系统中实际是VIM的别名\n\n**使用场景：**\n• 系统救援模式下的文件编辑\n• 资源受限的嵌入式系统\n• 快速编辑小文件\n• 远程服务器管理''',
                'code': '''# vi基础操作\n\n# 启动vi\nvi filename.txt          # 打开文件\nvi +10 file.txt         # 打开文件并跳到第10行\nvi +/pattern file.txt   # 打开文件并搜索pattern\n\n# 模式切换\ni                        # 进入插入模式（光标前）\na                        # 进入插入模式（光标后）\nA                        # 进入插入模式（行尾）\no                        # 新建行并进入插入模式\nO                        # 在上方新建行并进入插入模式\nEsc                      # 返回命令模式\n\n# 光标移动（命令模式）\nh                        # 左移一个字符\nj                        # 下移一行\nk                        # 上移一行\nl                        # 右移一个字符\nw                        # 移到下一个单词开头\nb                        # 移到上一个单词开头\n0                        # 移到行首\n$                        # 移到行尾\nG                        # 移到文件末尾\n1G 或 gg                 # 移到文件开头\n10G                      # 移到第10行\n\n# 文本编辑（命令模式）\nx                        # 删除当前字符\nX                        # 删除前一个字符\ndw                       # 删除单词\ndd                       # 删除整行\n3dd                      # 删除3行\nyy                       # 复制当前行\n3yy                      # 复制3行\np                        # 在光标后粘贴\nP                        # 在光标前粘贴\nu                        # 撤销上一个操作\n\n# 搜索和替换\n/pattern                 # 向下搜索\n?pattern                 # 向上搜索\nn                        # 下一个匹配\nN                        # 上一个匹配\n:s/old/new/              # 替换当前行第一个\n:s/old/new/g             # 替换当前行所有\n:%s/old/new/g            # 替换全文所有\n\n# 文件操作\n:w                       # 保存文件\n:w filename              # 另存为\n:q                       # 退出\n:q!                      # 强制退出不保存\n:wq                      # 保存并退出\n:x                       # 保存并退出（同:wq）\nZZ                       # 保存并退出（快捷键）''',
                'case': '''**案例：使用vi进行系统配置文件快速编辑**\n\n场景：在服务器维护时使用vi编辑关键配置文件\n\n```bash\n# 1. 编辑SSH配置文件\nsudo vi /etc/ssh/sshd_config\n\n# vi操作步骤：\n# - 按 /Port 搜索端口配置\n# - 按 n 找到下一个匹配\n# - 按 A 移到行尾并进入插入模式\n# - 修改端口号\n# - 按 Esc 返回命令模式\n# - 输入 :wq 保存并退出\n\n# 2. 编辑网络配置（CentOS）\nsudo vi /etc/sysconfig/network-scripts/ifcfg-eth0\n\n# 快速编辑技巧：\n# - 使用 dd 删除整行\n# - 使用 o 在下方新建行\n# - 使用 yy 和 p 复制粘贴行\n```\n\n**vi实用技巧：**\n\n1. **快速导航：**\n```bash\n# 跳转到特定行\n:50                      # 跳转到第50行\n50G                      # 同上\n\n# 快速移动\nCtrl+f                   # 向下翻页\nCtrl+b                   # 向上翻页\nCtrl+d                   # 向下半页\nCtrl+u                   # 向上半页\n```\n\n2. **批量操作：**\n```bash\n# 删除多行\n5dd                      # 删除5行\nd5d                      # 同上\n\n# 复制多行\n3yy                      # 复制3行\ny3y                      # 同上\n```\n\n3. **搜索技巧：**\n```bash\n# 搜索并替换\n:%s/192.168.1/10.0.0/g   # 替换IP地址\n:1,10s/old/new/g         # 只在1-10行替换\n```\n\n4. **紧急情况处理：**\n```bash\n# 如果vi异常退出，可能产生交换文件\n# 恢复文件：\nvi -r filename.txt\n\n# 删除交换文件：\nrm .filename.txt.swp\n```\n\n**系统管理中的vi应用：**\n\n1. **日志文件查看：**\n```bash\n# 查看系统日志\nsudo vi /var/log/messages\n\n# 快速定位到文件末尾\nG\n\n# 搜索错误信息\n/error\n```\n\n2. **配置文件备份和编辑：**\n```bash\n# 备份配置文件\ncp /etc/httpd/conf/httpd.conf /etc/httpd/conf/httpd.conf.bak\n\n# 编辑配置\nvi /etc/httpd/conf/httpd.conf\n```\n\n**vi与其他编辑器的选择：**\n• 紧急情况：优先使用vi（最基础，一定存在）\n• 日常开发：可选择VIM（功能更强大）\n• 简单编辑：可选择nano（更易上手）\n• 图形界面：可选择gedit、kate等'''
            },
            8: {
                'title': 'Nano编辑器入门与高效使用',
                'theory': '''Nano是一个简单易用的命令行文本编辑器，特别适合Linux初学者。它提供了直观的界面和简单的操作方式，是vi/vim的友好替代品。\n\n**Nano的特点：**\n• 用户友好，易于学习\n• 底部显示常用快捷键提示\n• 支持语法高亮\n• 支持多文件编辑\n• 占用资源少，启动快速\n• 支持鼠标操作（在支持的终端中）\n\n**Nano的优势：**\n• 学习曲线平缓，新手友好\n• 快捷键直观易记\n• 无需模式切换\n• 实时显示帮助信息\n• 支持搜索和替换\n• 支持拼写检查\n\n**适用场景：**\n• Linux初学者的首选编辑器\n• 快速编辑配置文件\n• 编写简单脚本\n• 查看和修改文本文件\n• 系统管理任务\n\n**与其他编辑器对比：**\n• 比vi/vim更容易上手\n• 比图形编辑器更轻量\n• 功能比vi丰富，比vim简单\n• 适合不需要复杂编辑功能的场景''',
                'code': '''# Nano基础操作\n\n# 启动nano\nnano filename.txt        # 打开文件\nnano +10 file.txt       # 打开文件并跳到第10行\nnano -w file.txt        # 禁用自动换行\nnano -T 4 file.txt      # 设置Tab宽度为4\n\n# 基本编辑操作（直接输入即可编辑）\n# 光标移动\nCtrl + A                 # 移到行首\nCtrl + E                 # 移到行尾\nCtrl + Y                 # 上一页\nCtrl + V                 # 下一页\nAlt + A                  # 选择文本（开始标记）\n\n# 文件操作\nCtrl + O                 # 保存文件（WriteOut）\nCtrl + X                 # 退出nano\nCtrl + R                 # 读取文件（插入另一个文件内容）\n\n# 编辑操作\nCtrl + K                 # 剪切当前行\nCtrl + U                 # 粘贴\nCtrl + 6                 # 开始选择文本\nAlt + 6                  # 复制选中文本\nCtrl + T                 # 拼写检查\n\n# 搜索和替换\nCtrl + W                 # 搜索\nCtrl + \\                 # 搜索并替换\nAlt + W                  # 重复上次搜索\nAlt + R                  # 替换\n\n# 导航操作\nCtrl + G                 # 显示帮助\nCtrl + C                 # 显示光标位置\nCtrl + _                 # 跳转到指定行号\nAlt + G                  # 跳转到指定行号\n\n# 其他有用操作\nCtrl + J                 # 对齐段落\nCtrl + L                 # 刷新屏幕\nAlt + U                  # 撤销\nAlt + E                  # 重做\nAlt + 3                  # 注释/取消注释\nAlt + }                  # 缩进\nAlt + {                  # 取消缩进''',
                'case': '''**案例：使用Nano进行日常系统管理**\n\n场景1：编辑系统配置文件\n\n```bash\n# 1. 编辑主机名\nsudo nano /etc/hostname\n\n# 操作步骤：\n# - 直接输入新的主机名\n# - 按 Ctrl+O 保存\n# - 按 Enter 确认文件名\n# - 按 Ctrl+X 退出\n\n# 2. 编辑hosts文件\nsudo nano /etc/hosts\n\n# 添加新的主机映射：\n# - 移动到文件末尾\n# - 输入：192.168.1.100 myserver.local\n# - 保存并退出\n```\n\n场景2：创建和编辑脚本文件\n\n```bash\n# 创建备份脚本\nnano backup_script.sh\n\n# 输入脚本内容：\n#!/bin/bash\n# 系统备份脚本\necho "开始备份..."\ntar -czf /backup/system_$(date +%Y%m%d).tar.gz /etc /home\necho "备份完成"\n\n# 保存并设置执行权限\nchmod +x backup_script.sh\n```\n\n**Nano高级技巧：**\n\n1. **多文件编辑：**\n```bash\n# 同时打开多个文件\nnano file1.txt file2.txt file3.txt\n\n# 在nano中切换文件\nAlt + <                  # 上一个文件\nAlt + >                  # 下一个文件\n```\n\n2. **搜索和替换技巧：**\n```bash\n# 在nano中进行搜索替换\n# 1. 按 Ctrl+\\ 进入替换模式\n# 2. 输入要搜索的文本\n# 3. 按 Enter\n# 4. 输入替换文本\n# 5. 按 Enter\n# 6. 选择替换选项：\n#    - Y: 替换当前匹配\n#    - N: 跳过当前匹配\n#    - A: 替换所有匹配\n```\n\n3. **配置nano：**\n```bash\n# 创建nano配置文件\nnano ~/.nanorc\n\n# 添加配置选项：\nset tabsize 4            # 设置Tab大小\nset autoindent           # 自动缩进\nset linenumbers          # 显示行号\nset mouse                # 启用鼠标支持\nset smooth               # 平滑滚动\nset softwrap             # 软换行\ninclude "/usr/share/nano/*.nanorc"  # 启用语法高亮\n```\n\n4. **语法高亮：**\n```bash\n# nano支持多种文件类型的语法高亮\n# 编辑不同类型的文件：\nnano script.py           # Python语法高亮\nnano config.json         # JSON语法高亮\nnano style.css           # CSS语法高亮\nnano index.html          # HTML语法高亮\n```\n\n**实用场景应用：**\n\n1. **快速编辑配置文件：**\n```bash\n# 编辑SSH配置\nsudo nano /etc/ssh/sshd_config\n\n# 编辑防火墙规则\nsudo nano /etc/iptables/rules.v4\n\n# 编辑定时任务\ncrontab -e  # 如果设置nano为默认编辑器\n```\n\n2. **日志文件查看和编辑：**\n```bash\n# 查看系统日志\nsudo nano /var/log/syslog\n\n# 使用 Ctrl+W 搜索特定错误\n# 使用 Ctrl+C 查看当前位置\n```\n\n3. **创建文档和说明文件：**\n```bash\n# 创建README文件\nnano README.md\n\n# 创建安装说明\nnano INSTALL.txt\n```\n\n**Nano vs 其他编辑器的选择建议：**\n\n• **选择nano的情况：**\n  - Linux新手\n  - 需要快速编辑文件\n  - 不需要复杂的编辑功能\n  - 偶尔使用命令行编辑器\n\n• **选择vi/vim的情况：**\n  - 需要高效的编辑操作\n  - 经常进行代码编辑\n  - 需要强大的搜索替换功能\n  - 系统资源极其有限\n\n• **nano的学习路径：**\n  1. 掌握基本的打开、编辑、保存、退出\n  2. 学习搜索和替换功能\n  3. 了解配置选项和语法高亮\n  4. 根据需要决定是否学习更高级的编辑器'''
            }
        },
        'Shell脚本基础语法与简单应用': {
            1: {
                'title': 'Shell脚本入门与环境配置',
                'theory': '''Shell脚本是Linux系统管理和自动化的重要工具。Shell作为命令行解释器，可以执行系统命令、处理文件和管理进程。常见的Shell包括bash、zsh、csh等，其中bash是最广泛使用的。\n\n**Shell脚本特点：**\n• 解释执行，无需编译\n• 语法简单，易于学习\n• 与系统命令紧密集成\n• 适合自动化任务\n\n**脚本基础：**\n• 脚本文件以#!/bin/bash开头（Shebang）\n• 文件需要执行权限：chmod +x script.sh\n• 执行方式：./script.sh 或 bash script.sh''',
                'code': '''#!/bin/bash\n# 第一个Shell脚本\necho "Hello, Shell World!"\n\n# 检查Shell类型\necho "当前Shell: $SHELL"\necho "Shell版本: $BASH_VERSION"\n\n# 脚本信息\necho "脚本名称: $0"\necho "执行时间: $(date)"\necho "当前用户: $USER"\necho "工作目录: $PWD"\n\n# 设置执行权限并运行\n# chmod +x hello.sh\n# ./hello.sh''',
                'case': '''**案例：系统信息收集脚本**\n\n创建一个收集系统基本信息的脚本：\n\n```bash\n#!/bin/bash\n# 文件名：system_info.sh\n\necho "=== 系统信息收集 ==="\necho "主机名: $(hostname)"\necho "操作系统: $(uname -s)"\necho "内核版本: $(uname -r)"\necho "系统架构: $(uname -m)"\necho "运行时间: $(uptime)"\necho "当前时间: $(date)"\necho "磁盘使用: $(df -h / | tail -1)"\necho "内存使用: $(free -h | grep Mem)"\n```\n\n**应用场景：**\n• 系统监控脚本\n• 环境检查工具\n• 自动化部署脚本'''
            },
            2: {
                'title': '变量定义与使用技巧',
                'theory': '''Shell变量是存储数据的容器，分为环境变量和用户自定义变量。变量赋值时等号两边不能有空格，使用时需要加$符号。\n\n**变量类型：**\n• 字符串变量：name="value"\n• 数值变量：count=10\n• 数组变量：arr=(a b c)\n• 只读变量：readonly PI=3.14\n\n**变量作用域：**\n• 局部变量：仅在当前Shell中有效\n• 环境变量：可被子进程继承\n• 全局变量：export导出的变量\n\n**特殊变量：**\n• $0：脚本名称\n• $1-$9：位置参数\n• $#：参数个数\n• $@：所有参数\n• $?：上一命令退出状态''',
                'code': '''#!/bin/bash\n# 变量定义与使用\n\n# 基本变量\nname="张三"\nage=25\npath="/home/user"\n\n# 变量使用\necho "姓名: $name"\necho "年龄: ${age}岁"\necho "路径: $path"\n\n# 命令替换\ncurrent_date=$(date +"%Y-%m-%d")\nfile_count=`ls | wc -l`\necho "当前日期: $current_date"\necho "文件数量: $file_count"\n\n# 数组变量\nfruits=("苹果" "香蕉" "橙子")\necho "第一个水果: ${fruits[0]}"\necho "所有水果: ${fruits[@]}"\necho "数组长度: ${#fruits[@]}"\n\n# 只读变量\nreadonly PI=3.14159\necho "圆周率: $PI"\n\n# 删除变量\nunset name\necho "删除后的name: $name"''',
                'case': '''**案例：配置文件管理脚本**\n\n创建一个管理应用配置的脚本：\n\n```bash\n#!/bin/bash\n# 文件名：config_manager.sh\n\n# 配置变量\nAPP_NAME="BigDataApp"\nAPP_VERSION="1.0.0"\nCONFIG_DIR="/etc/bigdata"\nLOG_DIR="/var/log/bigdata"\nDATA_DIR="/data/bigdata"\n\n# 数据库配置\nDB_HOST="localhost"\nDB_PORT=3306\nDB_NAME="bigdata_db"\nDB_USER="admin"\n\n# 创建配置文件\ncreate_config() {\n    echo "# $APP_NAME 配置文件" > "$CONFIG_DIR/app.conf"\n    echo "app.name=$APP_NAME" >> "$CONFIG_DIR/app.conf"\n    echo "app.version=$APP_VERSION" >> "$CONFIG_DIR/app.conf"\n    echo "db.host=$DB_HOST" >> "$CONFIG_DIR/app.conf"\n    echo "db.port=$DB_PORT" >> "$CONFIG_DIR/app.conf"\n    echo "配置文件已创建: $CONFIG_DIR/app.conf"\n}\n\n# 显示配置\nshow_config() {\n    echo "=== 应用配置 ==="\n    echo "应用名称: $APP_NAME"\n    echo "版本号: $APP_VERSION"\n    echo "配置目录: $CONFIG_DIR"\n    echo "日志目录: $LOG_DIR"\n    echo "数据目录: $DATA_DIR"\n}\n\n# 主程序\ncase $1 in\n    "create")\n        create_config\n        ;;\n    "show")\n        show_config\n        ;;\n    *)\n        echo "用法: $0 {create|show}"\n        ;;\nesac\n```'''
            },
            3: {
                'title': '条件判断与分支控制',
                'theory': '''Shell提供if-then-else条件判断结构，支持数值比较、字符串比较和文件测试。test命令或[]用于条件测试，&&和||用于逻辑运算。\n\n**比较操作符：**\n• 数值比较：-eq(等于)、-ne(不等于)、-gt(大于)、-lt(小于)\n• 字符串比较：=(等于)、!=(不等于)、-z(为空)、-n(非空)\n• 文件测试：-f(文件)、-d(目录)、-e(存在)、-r(可读)\n\n**逻辑操作符：**\n• &&：逻辑与\n• ||：逻辑或\n• !：逻辑非\n\n**case语句：**\n• 多分支选择结构\n• 支持模式匹配\n• 比多个if-elif更简洁''',
                'code': '''#!/bin/bash\n# 条件判断示例\n\n# 数值比较\nnum=10\nif [ $num -gt 5 ]; then\n    echo "数字大于5"\nelif [ $num -eq 5 ]; then\n    echo "数字等于5"\nelse\n    echo "数字小于5"\nfi\n\n# 字符串比较\nuser="admin"\nif [ "$user" = "admin" ]; then\n    echo "管理员用户"\nelse\n    echo "普通用户"\nfi\n\n# 文件测试\nfile="/etc/passwd"\nif [ -f "$file" ]; then\n    echo "文件存在"\n    if [ -r "$file" ]; then\n        echo "文件可读"\n    fi\nfi\n\n# case语句\necho "请选择操作: (start|stop|restart|status)"\nread action\ncase $action in\n    "start")\n        echo "启动服务..."\n        ;;\n    "stop")\n        echo "停止服务..."\n        ;;\n    "restart")\n        echo "重启服务..."\n        ;;\n    "status")\n        echo "查看状态..."\n        ;;\n    *)\n        echo "无效操作"\n        ;;\nesac''',
                'case': '''**案例：服务状态检查脚本**\n\n创建一个检查系统服务状态的脚本：\n\n```bash\n#!/bin/bash\n# 文件名：service_check.sh\n\n# 检查服务状态\ncheck_service() {\n    local service_name=$1\n    \n    if systemctl is-active --quiet "$service_name"; then\n        echo "✓ $service_name 服务正在运行"\n        return 0\n    else\n        echo "✗ $service_name 服务未运行"\n        return 1\n    fi\n}\n\n# 检查端口\ncheck_port() {\n    local port=$1\n    \n    if netstat -tuln | grep -q ":$port "; then\n        echo "✓ 端口 $port 正在监听"\n        return 0\n    else\n        echo "✗ 端口 $port 未监听"\n        return 1\n    fi\n}\n\n# 主检查逻辑\necho "=== 系统服务检查 ==="\n\n# 检查关键服务\nservices=("nginx" "mysql" "redis")\nfor service in "${services[@]}"; do\n    check_service "$service"\ndone\n\necho "\n=== 端口检查 ==="\n\n# 检查关键端口\nports=(80 443 3306 6379)\nfor port in "${ports[@]}"; do\n    check_port "$port"\ndone\n\n# 磁盘空间检查\necho "\n=== 磁盘空间检查 ==="\ndisk_usage=$(df / | tail -1 | awk '{print $5}' | sed 's/%//')\nif [ "$disk_usage" -gt 80 ]; then\n    echo "⚠ 磁盘使用率过高: ${disk_usage}%"\nelse\n    echo "✓ 磁盘使用率正常: ${disk_usage}%"\nfi\n```'''
            },
            4: {
                'title': '循环结构与流程控制',
                'theory': '''Shell支持for、while、until三种循环结构。for循环适合遍历列表，while循环适合条件循环，until循环在条件为假时执行。\n\n**for循环类型：**\n• 列表遍历：for item in list\n• 数值范围：for i in {1..10}\n• C风格：for ((i=0; i<10; i++))\n• 文件遍历：for file in *.txt\n\n**while循环：**\n• 条件为真时执行\n• 适合不确定次数的循环\n• 常用于读取文件\n\n**until循环：**\n• 条件为假时执行\n• 与while相反\n\n**流程控制：**\n• break：跳出循环\n• continue：跳过当前迭代\n• exit：退出脚本''',
                'code': '''#!/bin/bash\n# 循环结构示例\n\n# for循环 - 数值范围\necho "=== 数值循环 ==="\nfor i in {1..5}; do\n    echo "数字: $i"\ndone\n\n# for循环 - 列表遍历\necho "\n=== 列表遍历 ==="\nfruits=("苹果" "香蕉" "橙子")\nfor fruit in "${fruits[@]}"; do\n    echo "水果: $fruit"\ndone\n\n# for循环 - 文件遍历\necho "\n=== 文件遍历 ==="\nfor file in *.txt; do\n    if [ -f "$file" ]; then\n        echo "处理文件: $file"\n        wc -l "$file"\n    fi\ndone\n\n# while循环\necho "\n=== while循环 ==="\ncount=1\nwhile [ $count -le 3 ]; do\n    echo "计数: $count"\n    count=$((count + 1))\ndone\n\n# 读取文件\necho "\n=== 读取文件 ==="\nwhile IFS= read -r line; do\n    echo "行内容: $line"\ndone < "/etc/passwd" | head -5\n\n# until循环\necho "\n=== until循环 ==="\nnum=1\nuntil [ $num -gt 3 ]; do\n    echo "Until: $num"\n    num=$((num + 1))\ndone''',
                'case': '''**案例：批量文件处理脚本**\n\n创建一个批量处理日志文件的脚本：\n\n```bash\n#!/bin/bash\n# 文件名：batch_process.sh\n\n# 配置\nLOG_DIR="/var/log/app"\nARCHIVE_DIR="/var/log/archive"\nMAX_SIZE=100M  # 100MB\nDAYS_OLD=7\n\n# 创建归档目录\nmkdir -p "$ARCHIVE_DIR"\n\n# 处理大文件\necho "=== 处理大文件 ==="\nfor logfile in "$LOG_DIR"/*.log; do\n    if [ -f "$logfile" ]; then\n        # 检查文件大小\n        size=$(stat -f%z "$logfile" 2>/dev/null || stat -c%s "$logfile")\n        size_mb=$((size / 1024 / 1024))\n        \n        if [ $size_mb -gt 100 ]; then\n            echo "处理大文件: $logfile (${size_mb}MB)"\n            \n            # 压缩并移动\n            gzip "$logfile"\n            mv "${logfile}.gz" "$ARCHIVE_DIR/"\n            echo "已归档: ${logfile}.gz"\n        fi\n    fi\ndone\n\n# 清理旧文件\necho "\n=== 清理旧文件 ==="\nfind "$ARCHIVE_DIR" -name "*.gz" -mtime +$DAYS_OLD -type f | while read -r oldfile; do\n    echo "删除旧文件: $oldfile"\n    rm "$oldfile"\ndone\n\n# 统计处理结果\necho "\n=== 处理统计 ==="\nactive_logs=$(find "$LOG_DIR" -name "*.log" | wc -l)\narchived_logs=$(find "$ARCHIVE_DIR" -name "*.gz" | wc -l)\necho "活跃日志文件: $active_logs"\necho "归档日志文件: $archived_logs"\n\n# 磁盘使用情况\necho "\n=== 磁盘使用 ==="\ndu -sh "$LOG_DIR" "$ARCHIVE_DIR"\n```'''
            },
            5: {
                'title': '函数定义与模块化编程',
                'theory': '''Shell函数是可重用的代码块，提高脚本的模块化和可维护性。函数可以接收参数，通过$1、$2等访问参数，$#获取参数个数，$@获取所有参数。\n\n**函数特点：**\n• 代码重用，减少重复\n• 模块化设计，便于维护\n• 参数传递，灵活调用\n• 返回值，状态反馈\n\n**函数语法：**\n• 定义：function_name() { commands; }\n• 调用：function_name arg1 arg2\n• 局部变量：local var=value\n• 返回值：return 0-255\n\n**最佳实践：**\n• 函数名要有意义\n• 使用局部变量\n• 错误处理\n• 文档注释''',
                'code': '''#!/bin/bash\n# 函数定义与使用\n\n# 简单函数\ngreet() {\n    echo "Hello, $1!"\n}\n\n# 带参数检查的函数\ncheck_file() {\n    local filename=$1\n    \n    if [ $# -eq 0 ]; then\n        echo "错误: 请提供文件名"\n        return 1\n    fi\n    \n    if [ -f "$filename" ]; then\n        echo "文件 $filename 存在"\n        return 0\n    else\n        echo "文件 $filename 不存在"\n        return 1\n    fi\n}\n\n# 计算函数\ncalculate() {\n    local num1=$1\n    local num2=$2\n    local operation=$3\n    \n    case $operation in\n        "+")\n            echo $((num1 + num2))\n            ;;\n        "-")\n            echo $((num1 - num2))\n            ;;\n        "*")\n            echo $((num1 * num2))\n            ;;\n        "/")\n            if [ $num2 -ne 0 ]; then\n                echo $((num1 / num2))\n            else\n                echo "错误: 除数不能为0"\n                return 1\n            fi\n            ;;\n        *)\n            echo "错误: 不支持的操作"\n            return 1\n            ;;\n    esac\n}\n\n# 函数调用\ngreet "世界"\ncheck_file "/etc/passwd"\nresult=$(calculate 10 5 "+")\necho "计算结果: $result"''',
                'case': '''**案例：系统管理工具库**\n\n创建一个可重用的系统管理函数库：\n\n```bash\n#!/bin/bash\n# 文件名：syslib.sh - 系统管理函数库\n\n# 日志函数\nlog_info() {\n    echo "[$(date '+%Y-%m-%d %H:%M:%S')] INFO: $1"\n}\n\nlog_error() {\n    echo "[$(date '+%Y-%m-%d %H:%M:%S')] ERROR: $1" >&2\n}\n\n# 服务管理函数\nstart_service() {\n    local service_name=$1\n    \n    if [ -z "$service_name" ]; then\n        log_error "服务名不能为空"\n        return 1\n    fi\n    \n    log_info "启动服务: $service_name"\n    if systemctl start "$service_name"; then\n        log_info "服务 $service_name 启动成功"\n        return 0\n    else\n        log_error "服务 $service_name 启动失败"\n        return 1\n    fi\n}\n\n# 备份函数\nbackup_directory() {\n    local source_dir=$1\n    local backup_dir=$2\n    local timestamp=$(date +"%Y%m%d_%H%M%S")\n    \n    if [ ! -d "$source_dir" ]; then\n        log_error "源目录不存在: $source_dir"\n        return 1\n    fi\n    \n    mkdir -p "$backup_dir"\n    local backup_file="$backup_dir/backup_${timestamp}.tar.gz"\n    \n    log_info "开始备份: $source_dir -> $backup_file"\n    if tar -czf "$backup_file" -C "$(dirname "$source_dir")" "$(basename "$source_dir")"; then\n        log_info "备份完成: $backup_file"\n        return 0\n    else\n        log_error "备份失败"\n        return 1\n    fi\n}\n\n# 磁盘空间检查\ncheck_disk_space() {\n    local path=${1:-"/"}\n    local threshold=${2:-80}\n    \n    local usage=$(df "$path" | tail -1 | awk '{print $5}' | sed 's/%//')\n    \n    if [ "$usage" -gt "$threshold" ]; then\n        log_error "磁盘空间不足: $path 使用率 ${usage}% (阈值: ${threshold}%)"\n        return 1\n    else\n        log_info "磁盘空间正常: $path 使用率 ${usage}%"\n        return 0\n    fi\n}\n\n# 使用示例\n# source syslib.sh\n# start_service "nginx"\n# backup_directory "/etc" "/backup"\n# check_disk_space "/" 90\n```'''
            }
        },
        'Shell脚本进阶（文件操作与进程监控）': {
            1: {
                'title': '高级文件操作与文本处理',
                'theory': '''Shell脚本中的高级文件操作是系统管理的核心技能。''',
                'code': '''# 文件查找和处理\nfind /var/log -name "*.log" -mtime +7\ngrep "ERROR" /var/log/app.log | tail -10\nsed -i 's/old/new/g' config.txt''',
                'case': '''实际应用：日志文件管理和分析系统。'''
            },
            2: {
                'title': '进程监控与系统管理',
                'theory': '''进程监控是系统运维的重要组成部分。''',
                'code': '''# 进程监控\nps aux | grep nginx\ntop -p $(pgrep nginx)\nkill -9 $(pgrep defunct)''',
                'case': '''实际应用：自动化服务监控和重启。'''
            },
            3: {
                'title': '定时任务与作业调度',
                'theory': '''使用cron和at命令进行任务调度。''',
                'code': '''# 定时任务\ncrontab -e\n0 2 * * * /backup/script.sh\nat now + 1 hour''',
                'case': '''实际应用：自动化备份和维护任务。'''
            }
        },
        '基于Docker的Hadoop伪分布式集群搭建': {
            1: {
                'title': 'Hadoop集群架构设计',
                'theory': '''Hadoop分布式文件系统和计算框架。''',
                'code': '''# Hadoop配置\nexport HADOOP_HOME=/opt/hadoop\nexport PATH=$PATH:$HADOOP_HOME/bin''',
                'case': '''实际应用：大数据平台搭建。'''
            },
            2: {
                'title': 'HDFS文件系统配置',
                'theory': '''Hadoop分布式文件系统配置。''',
                'code': '''# HDFS命令\nhdfs namenode -format\nhdfs dfs -ls /\nhdfs dfs -put file.txt /''',
                'case': '''实际应用：大数据存储管理。'''
            },
            3: {
                'title': 'YARN资源管理器',
                'theory': '''YARN负责集群资源管理和作业调度。''',
                'code': '''# YARN配置\nyarn-daemon.sh start resourcemanager\nyarn-daemon.sh start nodemanager''',
                'case': '''实际应用：计算资源调度。'''
            },
            4: {
                'title': 'MapReduce作业执行',
                'theory': '''MapReduce分布式计算模型。''',
                'code': '''# MapReduce示例\nhadoop jar examples.jar wordcount input output\nhadoop fs -cat output/part-r-00000''',
                'case': '''实际应用：大数据批处理。'''
            },
            5: {
                'title': '集群监控与维护',
                'theory': '''Hadoop集群的监控和维护。''',
                'code': '''# 集群状态
hdfs dfsadmin -report
yarn node -list
hadoop fsck /''',
                'case': '''实际应用：生产环境运维。'''
            }
        },
        'Hadoop核心组件': {
            1: {
                'title': 'HDFS分布式文件系统详解',
                'theory': '''HDFS（Hadoop Distributed File System）是Hadoop生态系统的核心组件，是一个高容错、高吞吐量的分布式文件系统。

**HDFS核心概念：**
• NameNode：主节点，管理文件系统的命名空间和元数据
• DataNode：从节点，存储实际的数据块
• Block：数据块，HDFS将文件分割成固定大小的块（默认128MB）
• Replication：副本机制，每个数据块默认有3个副本

**HDFS架构特点：**
• 主从架构：一个NameNode + 多个DataNode
• 一次写入，多次读取：适合大文件存储
• 流式数据访问：适合批处理而非交互式应用
• 商用硬件：运行在普通硬件上，通过软件实现容错

**HDFS优势：**
• 高容错性：数据自动备份，节点故障自动恢复
• 高吞吐量：适合大数据量的批处理
• 可扩展性：可以扩展到数千个节点
• 成本效益：使用商用硬件降低成本

**HDFS适用场景：**
• 大数据存储和处理
• 数据仓库和数据湖
• 日志文件存储
• 机器学习数据集存储''',
                'code': '''# HDFS基础命令操作

# 查看HDFS文件系统
hdfs dfs -ls /                    # 列出根目录
hdfs dfs -ls /user               # 列出用户目录
hdfs dfs -ls -h /user            # 人性化显示文件大小

# 创建目录
hdfs dfs -mkdir /user/data       # 创建目录
hdfs dfs -mkdir -p /user/data/input  # 递归创建目录

# 文件上传
hdfs dfs -put localfile.txt /user/data/  # 上传文件
hdfs dfs -copyFromLocal local.txt /user/data/  # 从本地复制
hdfs dfs -moveFromLocal local.txt /user/data/  # 移动本地文件

# 文件下载
hdfs dfs -get /user/data/file.txt ./     # 下载文件
hdfs dfs -copyToLocal /user/data/file.txt ./  # 复制到本地

# 文件操作
hdfs dfs -cp /user/data/file1.txt /user/backup/  # 复制文件
hdfs dfs -mv /user/data/old.txt /user/data/new.txt  # 移动/重命名
hdfs dfs -rm /user/data/file.txt         # 删除文件
hdfs dfs -rm -r /user/data/olddir        # 递归删除目录

# 查看文件内容
hdfs dfs -cat /user/data/file.txt        # 查看文件内容
hdfs dfs -head /user/data/file.txt       # 查看文件头部
hdfs dfs -tail /user/data/file.txt       # 查看文件尾部

# 文件信息
hdfs dfs -du /user/data                  # 查看目录大小
hdfs dfs -du -h /user/data               # 人性化显示大小
hdfs dfs -count /user/data               # 统计文件和目录数量

# 权限管理
hdfs dfs -chmod 755 /user/data/file.txt  # 修改文件权限
hdfs dfs -chown user:group /user/data/   # 修改所有者

# 集群管理命令
hdfs dfsadmin -report                    # 查看集群状态
hdfs dfsadmin -safemode get              # 查看安全模式状态
hdfs dfsadmin -safemode leave            # 退出安全模式
hdfs fsck /                              # 文件系统检查
hdfs fsck / -files -blocks -locations   # 详细检查信息''',
                'case': '''**案例：构建企业级数据湖存储方案**

场景：为电商公司构建基于HDFS的数据湖，存储用户行为日志、交易数据和商品信息。

**1. 数据湖目录结构设计：**
```bash
# 创建数据湖目录结构
hdfs dfs -mkdir -p /datalake/raw/logs/user_behavior
hdfs dfs -mkdir -p /datalake/raw/logs/system
hdfs dfs -mkdir -p /datalake/raw/transaction
hdfs dfs -mkdir -p /datalake/raw/product
hdfs dfs -mkdir -p /datalake/processed/daily
hdfs dfs -mkdir -p /datalake/processed/monthly
hdfs dfs -mkdir -p /datalake/curated/reports
hdfs dfs -mkdir -p /datalake/curated/ml_datasets

# 设置目录权限
hdfs dfs -chmod 755 /datalake
hdfs dfs -chmod 750 /datalake/raw
hdfs dfs -chmod 755 /datalake/processed
hdfs dfs -chmod 755 /datalake/curated
```

**监控和运维要点：**
• 定期检查集群健康状态
• 监控存储空间使用率
• 备份NameNode元数据
• 设置数据备份和恢复策略
• 实施数据生命周期管理'''
            },
            2: {
                'title': 'YARN资源管理器深入解析',
                'theory': '''YARN（Yet Another Resource Negotiator）是Hadoop 2.0引入的资源管理系统，将资源管理和作业调度分离，提供了更好的可扩展性和多租户支持。

**YARN核心组件：**
• ResourceManager（RM）：全局资源管理器，负责整个集群的资源分配
• NodeManager（NM）：节点管理器，负责单个节点的资源管理
• ApplicationMaster（AM）：应用程序主控，负责单个应用的资源协调
• Container：资源容器，封装了CPU、内存等资源

**YARN工作流程：**
1. 客户端提交应用到ResourceManager
2. ResourceManager启动ApplicationMaster
3. ApplicationMaster向ResourceManager申请资源
4. ResourceManager分配Container给ApplicationMaster
5. ApplicationMaster在Container中启动任务
6. NodeManager监控Container的执行

**YARN调度器类型：**
• FIFO Scheduler：先进先出调度器，简单但不支持多租户
• Capacity Scheduler：容量调度器，支持多队列和资源保证
• Fair Scheduler：公平调度器，确保资源公平分配

**YARN优势：**
• 资源利用率高：支持多种计算框架共享集群
• 可扩展性强：支持数千个节点的大规模集群
• 多租户支持：不同用户和应用可以共享集群资源
• 容错性好：ApplicationMaster失败可以重启''',
                'code': '''# YARN集群管理命令

# 查看集群信息
yarn node -list                          # 查看所有节点
yarn node -status <nodeId>               # 查看特定节点状态
yarn rmadmin -printTopology              # 查看集群拓扑

# 应用程序管理
yarn application -list                   # 查看所有应用
yarn application -list -appStates RUNNING # 查看运行中的应用
yarn application -status <applicationId> # 查看应用状态
yarn application -kill <applicationId>   # 终止应用

# 队列管理
yarn queue -status <queueName>           # 查看队列状态
yarn rmadmin -refreshQueues              # 刷新队列配置

# 日志查看
yarn logs -applicationId <appId>         # 查看应用日志
yarn logs -applicationId <appId> -containerId <containerId>  # 查看容器日志

# 资源监控
yarn top                                 # 实时监控应用资源使用

# 配置管理
yarn rmadmin -refreshNodes               # 刷新节点配置
yarn rmadmin -refreshUserToGroupsMappings # 刷新用户组映射
yarn rmadmin -refreshSuperUserGroupsConfiguration # 刷新超级用户配置

# 提交MapReduce作业示例
hadoop jar $HADOOP_HOME/share/hadoop/mapreduce/hadoop-mapreduce-examples-*.jar \\
    wordcount \\
    -D mapreduce.job.queuename=default \\
    -D mapreduce.map.memory.mb=1024 \\
    -D mapreduce.reduce.memory.mb=2048 \\
    /input/data.txt /output/wordcount

# 提交Spark作业示例
spark-submit \\
    --class org.apache.spark.examples.SparkPi \\
    --master yarn \\
    --deploy-mode cluster \\
    --driver-memory 1g \\
    --executor-memory 1g \\
    --executor-cores 2 \\
    --queue production \\
    $SPARK_HOME/examples/jars/spark-examples_*.jar 10''',
                'case': '''**案例：企业级YARN多租户资源管理配置**

场景：为一家大数据公司配置YARN集群，支持开发、测试、生产三个环境的资源隔离和管理。

**1. 容量调度器配置（capacity-scheduler.xml）：**
```xml
<configuration>
    <!-- 根队列配置 -->
    <property>
        <name>yarn.scheduler.capacity.root.queues</name>
        <value>development,testing,production,default</value>
    </property>
    
    <!-- 开发队列配置 -->
    <property>
        <name>yarn.scheduler.capacity.root.development.capacity</name>
        <value>30</value>
    </property>
    <property>
        <name>yarn.scheduler.capacity.root.development.maximum-capacity</name>
        <value>50</value>
    </property>
</configuration>
```

**运维最佳实践：**
• 合理配置队列容量和最大容量
• 监控资源使用率，避免资源浪费
• 设置合适的Container大小
• 定期清理已完成的应用日志
• 实施资源配额管理'''
            },
            3: {
                'title': 'MapReduce计算框架实战应用',
                'theory': '''MapReduce是Hadoop的核心计算框架，采用分而治之的思想，将大数据处理任务分解为Map和Reduce两个阶段，实现大规模数据的并行处理。

**MapReduce核心概念：**
• Map阶段：将输入数据分片，并行处理，产生中间键值对
• Shuffle阶段：对Map输出进行排序、分组和传输
• Reduce阶段：对相同key的值进行聚合处理，产生最终结果
• InputFormat：定义如何读取输入数据
• OutputFormat：定义如何写入输出数据

**MapReduce工作流程：**
1. 输入分片：将大文件分割成多个输入分片
2. Map任务：每个分片启动一个Map任务并行处理
3. 分区排序：Map输出按key进行分区和排序
4. Shuffle传输：将Map输出传输给对应的Reduce任务
5. Reduce处理：对相同key的所有值进行聚合
6. 输出结果：将最终结果写入HDFS

**MapReduce优势：**
• 自动并行化：框架自动处理并行执行
• 容错性强：任务失败自动重试
• 可扩展性：支持数千个节点的大规模处理
• 数据本地性：优先在数据所在节点执行任务

**适用场景：**
• 大数据批处理
• 日志分析和挖掘
• 数据清洗和转换
• 统计分析和报表生成''',
                'code': '''# MapReduce作业提交和管理

# 基础作业提交
hadoop jar hadoop-mapreduce-examples.jar wordcount input output

# 带参数的作业提交
hadoop jar myapp.jar com.example.MyJob \\
    -D mapreduce.job.name="My Analysis Job" \\
    -D mapreduce.map.memory.mb=2048 \\
    -D mapreduce.reduce.memory.mb=4096 \\
    -D mapreduce.job.reduces=10 \\
    /input/data /output/result

# 作业监控
mapred job -list                         # 列出所有作业
mapred job -status <job_id>              # 查看作业状态
mapred job -counter <job_id> <group> <counter>  # 查看计数器
mapred job -kill <job_id>                # 终止作业

# 作业历史
mapred job -history <job_output_dir>     # 查看作业历史
yarn logs -applicationId <app_id>        # 查看应用日志

# 性能调优参数
-D mapreduce.map.memory.mb=2048          # Map任务内存
-D mapreduce.reduce.memory.mb=4096       # Reduce任务内存
-D mapreduce.map.java.opts="-Xmx1638m"   # Map JVM参数
-D mapreduce.reduce.java.opts="-Xmx3276m" # Reduce JVM参数
-D mapreduce.task.io.sort.mb=512         # 排序缓冲区大小
-D mapreduce.map.sort.spill.percent=0.8  # 溢写阈值
-D mapreduce.job.reduces=20              # Reduce任务数量
-D mapreduce.input.fileinputformat.split.maxsize=268435456  # 分片大小''',
                'case': '''**案例：电商网站用户行为分析系统**

场景：分析电商网站的用户访问日志，统计页面访问量、用户活跃度、商品热度等指标。

**1. 数据准备和预处理：**
```bash
#!/bin/bash
# 文件名：prepare_data.sh

# 创建输入目录
hdfs dfs -mkdir -p /ecommerce/logs/raw
hdfs dfs -mkdir -p /ecommerce/logs/processed

# 上传原始日志文件
hdfs dfs -put /var/log/nginx/access.log* /ecommerce/logs/raw/

# 数据清洗 - 过滤无效日志
hadoop jar log-cleaner.jar com.example.LogCleaner \\
    -D mapreduce.job.name="Log Cleaning" \\
    -D mapreduce.map.memory.mb=1024 \\
    /ecommerce/logs/raw /ecommerce/logs/processed

echo "数据预处理完成"
```

**业务价值和应用效果：**
• 实时了解网站访问热点和用户行为模式
• 为商品推荐和个性化营销提供数据支持
• 优化网站结构和用户体验
• 支持业务决策和运营策略制定'''
            }
        }
    }
    
    # 获取对应内容，如果没有则生成通用内容
    if module_name in content_map and topic_idx in content_map[module_name]:
        content = content_map[module_name][topic_idx]
        title = content['title']
        theory = content['theory']
        code = content['code']
        case = content['case']
    else:
        # 通用内容生成逻辑
        title = f"{module_name} · 知识点{topic_idx}"
        theory = f"本节将围绕{module_name}的第{topic_idx}个知识点展开，深入讲解核心概念和实际应用。"
        
        if 'Linux' in module_name:
            code = """# Linux基础命令\nls -la                    # 详细列表\npwd                       # 当前目录\ncd /path/to/dir          # 切换目录\nmkdir newdir             # 创建目录\ncp file1 file2           # 复制文件\nmv oldname newname       # 移动/重命名\nrm filename              # 删除文件"""
        elif 'Shell' in module_name:
            code = """#!/bin/bash\n# Shell脚本示例\n\n# 变量定义\nname="example"\ncount=10\n\n# 条件判断\nif [ $count -gt 5 ]; then\n    echo "Count is greater than 5"\nfi\n\n# 循环\nfor i in {1..5}; do\n    echo "Number: $i"\ndone"""
        elif 'Docker' in module_name:
            code = """# Docker基础命令\ndocker pull ubuntu:20.04\ndocker run -it ubuntu:20.04 bash\ndocker ps                 # 查看运行容器\ndocker images            # 查看镜像\n\n# Dockerfile示例\nFROM ubuntu:20.04\nRUN apt-get update\nCOPY . /app\nWORKDIR /app\nCMD [\"python\", \"app.py\"]"""
        else:
            code = f"# {module_name} 相关代码示例\necho \"学习 {module_name} 第 {topic_idx} 个知识点\""
        
        case = f"**实际案例：**\n\n在{module_name}的学习中，第{topic_idx}个知识点的典型应用场景包括系统管理、自动化运维和大数据处理等领域。通过实际操作可以更好地理解和掌握相关技能。"
    
    # 生成练习题
    exercises = [
        f"**理论练习：** 请用自己的话总结本节的核心概念，并说明在实际工作中的应用价值。",
        f"**实践练习：** 根据提供的代码示例，完成一个相关的小项目或任务。",
        f"**扩展练习：** 查阅相关资料，了解本知识点的高级用法或最佳实践。"
    ]
    
    return {
        'title': title,
        'theory': theory,
        'code': code,
        'case': case,
        'exercises': exercises,
        'images': random.sample(IMAGES, 2)
    }


def upsert_module_with_topics(c, ord_num, title, desc, topic_count=5):
    # Linux基础操作模块有8个知识点（包含VIM、vi、Nano）
    if ord_num == 1:  # Linux基础操作与环境熟悉
        topic_count = 8
    # Hadoop核心组件模块有3个知识点（HDFS、YARN、MapReduce）
    elif ord_num == 6:  # Hadoop核心组件
        topic_count = 3
    row = c.execute('SELECT id FROM modules WHERE ord=?', (ord_num,)).fetchone()
    if not row:
        cur = c.execute('INSERT INTO modules(title, description, ord) VALUES(?,?,?)', (title, desc, ord_num))
        mid = cur.lastrowid
    else:
        mid = row['id']
    # topics
    for t in range(1, topic_count + 1):
        r2 = c.execute('SELECT id FROM topics WHERE module_id=? AND ord=?', (mid, t)).fetchone()
        if not r2:
            # 生成内容并获取具体标题
            data = gen_content(title, t)
            ttitle = data.get('title', f"知识点 {t}")
            cur_t = c.execute('INSERT INTO topics(module_id,title,ord) VALUES(?,?,?)', (mid, ttitle, t))
            tid = cur_t.lastrowid
            c.execute('INSERT INTO contents(topic_id,data) VALUES(?,?)', (tid, json.dumps(data, ensure_ascii=False)))


def seed():
    with get_db() as c:
        # modules/topics/contents upsert
        for ord_num, mtitle, mdesc in MODULES_META:
            upsert_module_with_topics(c, ord_num, mtitle, mdesc, topic_count=5)
        # exams 5 sets * 30 (if not exists)
        for s in range(1, 6):
            ex = c.execute('SELECT id FROM exam_sets WHERE name=?', (f'模拟考试第{s}套',)).fetchone()
            if ex:
                continue
            cur_es = c.execute('INSERT INTO exam_sets(name) VALUES(?)', (f'模拟考试第{s}套',))
            exam_id = cur_es.lastrowid
            ordv = 1
            # for knowledge_ref mapping
            module_count = len(MODULES_META)
            # MCQ 12
            for i in range(12):
                prompt = f"[选择题] 第{s}套 第{i + 1}题：Linux命令ls的作用是？"
                opts = ['列出目录内容', '删除文件', '显示文件内容', '切换目录']
                ans = 'A'
                mod_ord = ((ordv - 1) % module_count) + 1
                topic_ord = ((ordv - 1) % 5) + 1
                mid = c.execute('SELECT id FROM modules WHERE ord=?', (mod_ord,)).fetchone()['id']
                c.execute('INSERT INTO questions(exam_id,qtype,prompt,options,answer,score,ord,knowledge_ref) VALUES(?,?,?,?,?,?,?,?)',
                          (exam_id, 'mcq', prompt, json.dumps(opts, ensure_ascii=False), ans, 2, ordv, f"{mid}:{topic_ord}"))
                ordv += 1
            # TF 8
            for i in range(8):
                prompt = f"[判断题] 第{s}套 第{i + 1}题：/root为普通用户家目录。"
                mod_ord = ((ordv - 1) % module_count) + 1
                topic_ord = ((ordv - 1) % 5) + 1
                mid = c.execute('SELECT id FROM modules WHERE ord=?', (mod_ord,)).fetchone()['id']
                c.execute('INSERT INTO questions(exam_id,qtype,prompt,options,answer,score,ord,knowledge_ref) VALUES(?,?,?,?,?,?,?,?)',
                          (exam_id, 'tf', prompt, None, 'False', 2, ordv, f"{mid}:{topic_ord}"))
                ordv += 1
            # Fill 6
            for i in range(6):
                prompt = f"[填空题] 第{s}套 第{i + 1}题：查看当前路径的命令是____。"
                mod_ord = ((ordv - 1) % module_count) + 1
                topic_ord = ((ordv - 1) % 5) + 1
                mid = c.execute('SELECT id FROM modules WHERE ord=?', (mod_ord,)).fetchone()['id']
                c.execute('INSERT INTO questions(exam_id,qtype,prompt,options,answer,score,ord,knowledge_ref) VALUES(?,?,?,?,?,?,?,?)',
                          (exam_id, 'fill', prompt, None, 'pwd', 2, ordv, f"{mid}:{topic_ord}"))
                ordv += 1
            # Short 4
            for i in range(4):
                prompt = f"[简答题] 第{s}套 第{i + 1}题：简述使用Docker部署Hadoop伪分布式的大致步骤。"
                ans = json.dumps({'keywords': ['拉取镜像', '配置core-site', '启动容器', 'HDFS格式化', 'NameNode/DataNode']}, ensure_ascii=False)
                mod_ord = ((ordv - 1) % module_count) + 1
                topic_ord = ((ordv - 1) % 5) + 1
                mid = c.execute('SELECT id FROM modules WHERE ord=?', (mod_ord,)).fetchone()['id']
                c.execute('INSERT INTO questions(exam_id,qtype,prompt,options,answer,score,ord,knowledge_ref) VALUES(?,?,?,?,?,?,?,?)',
                          (exam_id, 'short', prompt, None, ans, 4, ordv, f"{mid}:{topic_ord}"))
                ordv += 1
        
        # 为Hadoop核心组件模块添加专门的练习题
        hadoop_module_id = c.execute('SELECT id FROM modules WHERE ord=?', (6,)).fetchone()
        if hadoop_module_id:
            hadoop_module_id = hadoop_module_id['id']
            
            # 创建Hadoop专项练习考试集
            hadoop_exam = c.execute('SELECT id FROM exam_sets WHERE name=?', ('Hadoop核心组件专项练习',)).fetchone()
            if not hadoop_exam:
                cur_hadoop = c.execute('INSERT INTO exam_sets(name) VALUES(?)', ('Hadoop核心组件专项练习',))
                hadoop_exam_id = cur_hadoop.lastrowid
                ordv = 1
                
                # HDFS相关题目
                hdfs_questions = [
                    # 选择题
                    {'type': 'mcq', 'prompt': 'HDFS中NameNode的主要作用是什么？', 
                     'options': ['存储文件数据块', '管理文件系统命名空间和元数据', '执行MapReduce任务', '监控集群状态'], 
                     'answer': 'B', 'topic': 1},
                    {'type': 'mcq', 'prompt': 'HDFS默认的数据块大小是多少？', 
                     'options': ['64MB', '128MB', '256MB', '512MB'], 
                     'answer': 'B', 'topic': 1},
                    # 判断题
                    {'type': 'tf', 'prompt': 'HDFS支持随机写入操作。', 
                     'answer': 'False', 'topic': 1},
                    {'type': 'tf', 'prompt': 'DataNode负责存储实际的文件数据块。', 
                     'answer': 'True', 'topic': 1},
                    # 填空题
                    {'type': 'fill', 'prompt': 'HDFS中负责管理文件系统命名空间的组件是____。', 
                     'answer': 'NameNode', 'topic': 1},
                    # 简答题
                    {'type': 'short', 'prompt': '简述HDFS的读写流程。', 
                     'answer': {'keywords': ['客户端请求', 'NameNode返回DataNode列表', '直接与DataNode通信', '数据块复制', '确认写入']}, 'topic': 1}
                ]
                
                # YARN相关题目
                yarn_questions = [
                    # 选择题
                    {'type': 'mcq', 'prompt': 'YARN中ResourceManager的主要职责是什么？', 
                     'options': ['执行具体任务', '管理集群资源分配', '存储数据', '监控网络状态'], 
                     'answer': 'B', 'topic': 2},
                    {'type': 'mcq', 'prompt': 'ApplicationMaster的作用是什么？', 
                     'options': ['管理整个集群', '协调单个应用程序的执行', '存储应用数据', '监控硬件状态'], 
                     'answer': 'B', 'topic': 2},
                    # 判断题
                    {'type': 'tf', 'prompt': 'NodeManager运行在集群的每个工作节点上。', 
                     'answer': 'True', 'topic': 2},
                    {'type': 'tf', 'prompt': 'YARN只能运行MapReduce应用程序。', 
                     'answer': 'False', 'topic': 2},
                    # 填空题
                    {'type': 'fill', 'prompt': 'YARN中负责单个应用程序资源协调的组件是____。', 
                     'answer': 'ApplicationMaster', 'topic': 2},
                    # 简答题
                    {'type': 'short', 'prompt': '描述YARN的资源调度流程。', 
                     'answer': {'keywords': ['客户端提交应用', 'ResourceManager分配ApplicationMaster', 'ApplicationMaster请求资源', 'NodeManager启动Container', '任务执行']}, 'topic': 2}
                ]
                
                # MapReduce相关题目
                mapreduce_questions = [
                    # 选择题
                    {'type': 'mcq', 'prompt': 'MapReduce中Map阶段的主要作用是什么？', 
                     'options': ['汇总结果', '处理和转换输入数据', '存储数据', '监控任务'], 
                     'answer': 'B', 'topic': 3},
                    {'type': 'mcq', 'prompt': 'Shuffle阶段发生在哪两个阶段之间？', 
                     'options': ['输入和Map', 'Map和Reduce', 'Reduce和输出', '输入和输出'], 
                     'answer': 'B', 'topic': 3},
                    # 判断题
                    {'type': 'tf', 'prompt': 'MapReduce程序必须同时包含Map和Reduce阶段。', 
                     'answer': 'False', 'topic': 3},
                    {'type': 'tf', 'prompt': 'Combiner可以减少网络传输的数据量。', 
                     'answer': 'True', 'topic': 3},
                    # 填空题
                    {'type': 'fill', 'prompt': 'MapReduce中负责汇总Map输出结果的阶段是____。', 
                     'answer': 'Reduce', 'topic': 3},
                    # 简答题
                    {'type': 'short', 'prompt': '解释MapReduce的工作原理和优势。', 
                     'answer': {'keywords': ['分而治之', '并行处理', '容错性', 'Map映射', 'Reduce归约', '可扩展性']}, 'topic': 3}
                ]
                
                # 插入所有题目
                all_questions = hdfs_questions + yarn_questions + mapreduce_questions
                for q in all_questions:
                    if q['type'] == 'mcq':
                        c.execute('INSERT INTO questions(exam_id,qtype,prompt,options,answer,score,ord,knowledge_ref) VALUES(?,?,?,?,?,?,?,?)',
                                  (hadoop_exam_id, 'mcq', q['prompt'], json.dumps(q['options'], ensure_ascii=False), q['answer'], 2, ordv, f"{hadoop_module_id}:{q['topic']}"))
                    elif q['type'] == 'tf':
                        c.execute('INSERT INTO questions(exam_id,qtype,prompt,options,answer,score,ord,knowledge_ref) VALUES(?,?,?,?,?,?,?,?)',
                                  (hadoop_exam_id, 'tf', q['prompt'], None, q['answer'], 2, ordv, f"{hadoop_module_id}:{q['topic']}"))
                    elif q['type'] == 'fill':
                        c.execute('INSERT INTO questions(exam_id,qtype,prompt,options,answer,score,ord,knowledge_ref) VALUES(?,?,?,?,?,?,?,?)',
                                  (hadoop_exam_id, 'fill', q['prompt'], None, q['answer'], 2, ordv, f"{hadoop_module_id}:{q['topic']}"))
                    elif q['type'] == 'short':
                        c.execute('INSERT INTO questions(exam_id,qtype,prompt,options,answer,score,ord,knowledge_ref) VALUES(?,?,?,?,?,?,?,?)',
                                  (hadoop_exam_id, 'short', q['prompt'], None, json.dumps(q['answer'], ensure_ascii=False), 4, ordv, f"{hadoop_module_id}:{q['topic']}"))
                    ordv += 1

# ---------- Auth helpers ----------

def make_token():
    return secrets.token_hex(32)


def get_user_by_token(token):
    if not token:
        return None
    with get_db() as c:
        row = c.execute('SELECT u.* FROM tokens t JOIN users u ON t.user_id=u.id WHERE t.token=?', (token,)).fetchone()
        return dict(row) if row else None


def auth_required(role=None):
    def deco(fn):
        def wrapper(*args, **kwargs):
            auth = request.headers.get('Authorization', '')
            token = auth.split('Bearer ')[-1] if 'Bearer ' in auth else auth
            user = get_user_by_token(token)
            if not user:
                return jsonify({'error': 'unauthorized'}), 401
            if role and user.get('role') != role:
                return jsonify({'error': 'forbidden'}), 403
            request.user = user
            return fn(*args, **kwargs)
        wrapper.__name__ = fn.__name__
        return wrapper
    return deco

# ---------- API: content ----------
@app.route('/api/modules')
def api_modules():
    with get_db() as c:
        rows = c.execute('''
            SELECT m.id, m.title, m.description, m.ord, COUNT(t.id) as topics_count
            FROM modules m
            LEFT JOIN topics t ON m.id = t.module_id
            GROUP BY m.id, m.title, m.description, m.ord
            ORDER BY m.ord
        ''').fetchall()
        
        # 转换为字典并添加topics字段
        modules = []
        for row in rows:
            module_dict = dict(row)
            # 创建一个模拟的topics数组，长度等于topics_count
            module_dict['topics'] = [{'id': i} for i in range(module_dict['topics_count'])]
            modules.append(module_dict)
        
        return jsonify(modules)


@app.route('/api/modules/<int:mid>/topics')
def api_topics(mid):
    with get_db() as c:
        rows = c.execute('SELECT id,title,ord FROM topics WHERE module_id=? ORDER BY ord', (mid,)).fetchall()
        return jsonify([dict(r) for r in rows])


@app.route('/api/topics/<int:tid>/content')
def api_content(tid):
    with get_db() as c:
        row = c.execute('SELECT data FROM contents WHERE topic_id=?', (tid,)).fetchone()
        return jsonify(json.loads(row['data']) if row else {})

# ---------- API: exams ----------
@app.route('/api/exams')
def api_exams():
    with get_db() as c:
        rows = c.execute('SELECT id,name FROM exam_sets').fetchall()
        return jsonify([dict(r) for r in rows])


@app.route('/api/exams/<int:eid>')
def api_exam_detail(eid):
    with get_db() as c:
        rows = c.execute('SELECT id,qtype,prompt,options,score,ord FROM questions WHERE exam_id=? ORDER BY ord', (eid,)).fetchall()
        out = []
        for r in rows:
            opts = json.loads(r['options']) if r['options'] else None
            out.append({'id': r['id'], 'type': r['qtype'], 'prompt': r['prompt'], 'options': opts, 'score': r['score'], 'ord': r['ord']})
        return jsonify(out)


@app.route('/api/exams/<int:eid>/submit', methods=['POST'])
def api_exam_submit(eid):
    payload = request.json or {}
    answers = payload.get('answers', {})  # {qid: user_answer}
    total = 0
    got = 0
    detail = []
    wrong_qids = []
    with get_db() as c:
        rows = c.execute('SELECT id,qtype,answer,score,knowledge_ref FROM questions WHERE exam_id=?', (eid,)).fetchall()
        for r in rows:
            qid, qtype, ans, score, kref = r['id'], r['qtype'], r['answer'], r['score'], r['knowledge_ref']
            ua = answers.get(str(qid)) or answers.get(qid)
            correct = False
            part = 0.0
            if qtype == 'mcq':
                correct = (str(ua).strip().upper() == ans.upper())
            elif qtype == 'tf':
                correct = (str(ua).strip().lower() in ['true', 't', '1', '是', '对'] and ans == 'True') or \
                          (str(ua).strip().lower() in ['false', 'f', '0', '否', '错'] and ans == 'False')
            elif qtype == 'fill':
                correct = (str(ua).strip().lower() == ans.lower())
            else:  # short
                try:
                    kw = json.loads(ans).get('keywords', [])
                except Exception:
                    kw = []
                hit = sum(1 for k in kw if k.lower() in str(ua or '').lower())
                if kw:
                    part = min(1.0, hit / len(kw))
                correct = part > 0.6
            s = score if correct else (int(score * part) if part > 0 else 0)
            if s < score:
                wrong_qids.append(qid)
            got += s
            total += score
            detail.append({'qid': qid, 'type': qtype, 'score': s, 'max': score, 'kref': kref})
        # build suggestions by knowledge_ref
        kmap = {}
        for q in detail:
            if q['qid'] in wrong_qids and q['kref']:
                try:
                    mid, tord = map(int, str(q['kref']).split(':'))
                    row_m = c.execute('SELECT title FROM modules WHERE id=?', (mid,)).fetchone()
                    row_t = c.execute('SELECT title FROM topics WHERE module_id=? AND ord=?', (mid, tord)).fetchone()
                    key = f"{row_m['title']} - {row_t['title']}" if row_m and row_t else '未映射知识点'
                    kmap.setdefault(key, 0)
                    kmap[key] += 1
                except Exception:
                    pass
        suggestions = [f"优先复习：{k}（错题数 {v}），回到该模块阅读理论与案例，并完成练习。" for k, v in sorted(kmap.items(), key=lambda x: -x[1])]
    res = {'score': got, 'total': total, 'rate': round(got / total * 100, 2) if total else 0, 'detail': detail, 'wrong_qids': wrong_qids, 'suggestions': suggestions}

    # save if user logged in
    auth = request.headers.get('Authorization', '')
    token = auth.split('Bearer ')[-1] if 'Bearer ' in auth else auth
    user = get_user_by_token(token)
    if user:
        with get_db() as c:
            c.execute('INSERT INTO submissions(user_id,exam_id,score,total,rate,detail,wrong_qids,suggestions) VALUES(?,?,?,?,?,?,?,?)',
                      (user['id'], eid, got, total, res['rate'], json.dumps(detail, ensure_ascii=False), json.dumps(wrong_qids), json.dumps(suggestions, ensure_ascii=False)))
    return jsonify(res)

# ---------- API: auth & users ----------
@app.route('/api/auth/login', methods=['POST'])
def api_login():
    data = request.json or {}
    sid = data.get('student_id')
    username = data.get('username')
    password = data.get('password', '')
    with get_db() as c:
        row = None
        if sid:
            row = c.execute('SELECT * FROM users WHERE student_id=?', (sid,)).fetchone()
        if not row and username:
            row = c.execute('SELECT * FROM users WHERE username=?', (username,)).fetchone()
        if not row or not check_password_hash(row['password_hash'], password):
            return jsonify({'error': '账号或密码不正确'}), 400
        token = make_token()
        c.execute('INSERT INTO tokens(user_id, token) VALUES(?,?)', (row['id'], token))
        return jsonify({'token': token, 'user': {'id': row['id'], 'name': row['name'], 'username': row['username'], 'student_id': row['student_id'], 'role': row['role']}})


@app.route('/api/auth/register', methods=['POST'])
def api_register():
    data = request.json or {}
    sid = data.get('student_id')
    name = data.get('name')
    password = data.get('password', '123456')
    if not sid or not name:
        return jsonify({'error': '学号与姓名必填'}), 400
    with get_db() as c:
        try:
            c.execute('INSERT INTO users(student_id,name,role,password_hash) VALUES(?,?,?,?)', (sid, name, 'student', generate_password_hash(password)))
            return jsonify({'ok': True})
        except Exception:
            return jsonify({'error': '学号已存在'}), 400


@app.route('/api/auth/me')
@auth_required()
def api_me():
    u = request.user
    return jsonify({'id': u['id'], 'name': u['name'], 'username': u['username'], 'student_id': u['student_id'], 'role': u['role']})

# ---------- API: history ----------
@app.route('/api/my/scores')
@auth_required()
def api_my_scores():
    u = request.user
    with get_db() as c:
        rows = c.execute('''
            SELECT s.id, s.exam_id, e.name as exam_name, s.score, s.total, s.rate, s.created_at
            FROM submissions s JOIN exam_sets e ON s.exam_id=e.id
            WHERE s.user_id=? ORDER BY s.id DESC
        ''', (u['id'],)).fetchall()
        return jsonify([dict(r) for r in rows])


@app.route('/api/my/scores/<int:sid>')
@auth_required()
def api_my_score_detail(sid):
    u = request.user
    with get_db() as c:
        r = c.execute('''
            SELECT s.*, e.name as exam_name FROM submissions s JOIN exam_sets e ON s.exam_id=e.id
            WHERE s.id=? AND s.user_id=?
        ''', (sid, u['id'])).fetchone()
        if not r:
            return jsonify({}), 404
        detail = json.loads(r['detail']) if r['detail'] else []
        wrong_qids = json.loads(r['wrong_qids']) if r['wrong_qids'] else []
        suggestions = json.loads(r['suggestions']) if r['suggestions'] else []
        # enrich wrong questions
        wrongs = []
        for qid in wrong_qids:
            q = c.execute('SELECT id,prompt,knowledge_ref FROM questions WHERE id=?', (qid,)).fetchone()
            mod_title = topic_title = None
            if q and q['knowledge_ref']:
                try:
                    mid, tord = map(int, q['knowledge_ref'].split(':'))
                    m = c.execute('SELECT title FROM modules WHERE id=?', (mid,)).fetchone()
                    t = c.execute('SELECT title FROM topics WHERE module_id=? AND ord=?', (mid, tord)).fetchone()
                    mod_title = m['title'] if m else None
                    topic_title = t['title'] if t else None
                except Exception:
                    pass
            wrongs.append({'id': qid, 'prompt': q['prompt'] if q else '', 'module': mod_title, 'topic': topic_title})
        return jsonify({'id': r['id'], 'exam_name': r['exam_name'], 'score': r['score'], 'total': r['total'], 'rate': r['rate'], 'created_at': r['created_at'], 'wrongs': wrongs, 'suggestions': suggestions, 'detail': detail})

# ---------- API: admin import ----------
@app.route('/api/admin/import_students', methods=['POST'])
@auth_required(role='admin')
def api_admin_import():
    if 'file' not in request.files:
        return jsonify({'error': '缺少文件'}), 400
    f = request.files['file']
    from openpyxl import load_workbook
    wb = load_workbook(filename=f, data_only=True)
    ws = wb.active
    # expect headers like: 学号, 姓名
    count = 0
    with get_db() as c:
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue
            sid = str(row[0]).strip() if row and row[0] is not None else None
            name = str(row[1]).strip() if row and row[1] is not None else None
            if not sid or not name:
                continue
            try:
                c.execute('INSERT OR IGNORE INTO users(student_id,name,role,password_hash) VALUES(?,?,?,?)', (sid, name, 'student', generate_password_hash('123456')))
                # update name if exists
                c.execute('UPDATE users SET name=? WHERE student_id=?', (name, sid))
                count += 1
            except Exception:
                pass
    return jsonify({'ok': True, 'imported': count})

# ---------- API: LLM generation ----------
@app.route('/api/generate/topic/<int:tid>', methods=['POST'])
def api_generate_topic(tid):
    # fetch context
    with get_db() as c:
        r = c.execute('SELECT t.id as tid, t.ord as tord, m.id as mid, m.title as mtitle FROM topics t JOIN modules m ON t.module_id=m.id WHERE t.id=?', (tid,)).fetchone()
    if not r:
        return jsonify({'error': 'topic not found'}), 404
    api_key = os.environ.get('OPENAI_API_KEY')
    if not api_key or OpenAI is None:
        # fallback to local generator
        data = gen_content(r['mtitle'], r['tord'])
    else:
        client = OpenAI(api_key=api_key)
        prompt = f"你是资深大数据讲师。面向零基础学员，用通俗易懂的语言讲解模块《{r['mtitle']}》中‘知识点{r['tord']}’的内容。输出分为：1) 理论讲解（分点+比喻），2) 示例代码（含注释），3) 案例分析（一步步推演），4) 练习（3题），5) 小结（3条）。用中文，适当加入实践建议。"
        try:
            comp = client.chat.completions.create(
                model='gpt-4o-mini',
                messages=[{'role': 'system', 'content': '你是严谨而耐心的教学助理。'}, {'role': 'user', 'content': prompt}],
                temperature=0.3
            )
            txt = comp.choices[0].message.content
        except Exception as e:
            txt = None
        data = gen_content(r['mtitle'], r['tord'])
        if txt:
            data['theory'] = txt
    # save
    with get_db() as c:
        c.execute('UPDATE contents SET data=? WHERE topic_id=?', (json.dumps(data, ensure_ascii=False), tid))
    return jsonify(data)

# ---------- Main ----------
if __name__ == '__main__':
    init_db()
    seed()
    app.run(host='0.0.0.0', port=90, debug=False)